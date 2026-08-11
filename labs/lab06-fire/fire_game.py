"""
╔══════════════════════════════════════════════════════════════════╗
║   ОГНЕННЫЙ ФРОНТ: Защита тайги — v2 «Лаборатория»               abelme ║
║   Научно-образовательная стратегия для 8–9 классов               ║
║                                                                  ║
║   Научная модель: клеточный автомат Бака–Ченя–Тангена (1990)     ║
║   с расширением на ветер, влажность и тип горючего               ║
║                                                                  ║
║   ИЗМЕНЕНИЯ v2:                                                  ║
║     • Лаборатория перколяции (отдельный режим из меню)           ║
║     • Встроенная панель экспериментов в основной игре            ║
║     • Параметризация карты (плотность, тип, бюджет)              ║
║     • Визуализация вероятностей при наведении на огонь           ║
║     • Слоу-мо режим с пошаговым показом расчёта                  ║
║     • Автопрогон 100 симуляций (быстро/с графикой)               ║
║     • Журнал экспериментов сохраняется между запусками           ║
║     • Экспорт лаборатории в Excel                                ║
║     • Фикс бага с типом леса при тушении (forest_original)       ║
║                                                                  ║
║   Запуск:     python fire_game.py                                ║
║   Сборка:     pyinstaller --onefile --windowed fire_game.py      ║
║   Зависимости: pip install pygame openpyxl                       ║
║                                                                  ║
║   Управление:                                                    ║
║     ЛКМ   — поставить инструмент / выбрать юнит / цель           ║
║     ПКМ   — снять выбор / панорамирование (с зажатием)           ║
║     ESC   — отмена/меню                                          ║
║     SPACE — следующий день (в фазе сезона)                       ║
║     ENTER — начать сезон (в фазе подготовки)                     ║
║     S     — режим слоу-мо on/off                                 ║
║     A     — автопроигрывание (в лаборатории)                     ║
║     P     — пауза автопрогона                                    ║
║     +/-   — зум                                                  ║
║     Home  — сброс камеры                                         ║
╚══════════════════════════════════════════════════════════════════╝
"""

import pygame
import random
import math
import sys
import os
import json
from datetime import datetime
from collections import deque

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

pygame.init()
try:
    pygame.mixer.init()
except Exception:
    pass


# ═══════════════════════════════════════════════════════════════════
#   ЧАСТЬ 1: ЯДРО (константы, модель, игровое состояние)
# ═══════════════════════════════════════════════════════════════════

SCREEN_W = 1600
SCREEN_H = 900

GRID_W = 60
GRID_H = 40

TILE_W_BASE = 20
TILE_H_BASE = 10

TILE_W = TILE_W_BASE
TILE_H = TILE_H_BASE

CAM_ZOOM = 1.0
CAM_ZOOM_MIN = 0.6
CAM_ZOOM_MAX = 2.5
CAM_PAN_X = 0
CAM_PAN_Y = 0

MAP_AREA_TOP = 60
MAP_AREA_BOTTOM = SCREEN_H - 160
MAP_AREA_LEFT = 0
MAP_AREA_RIGHT = SCREEN_W - 320

MAP_AREA_W = MAP_AREA_RIGHT - MAP_AREA_LEFT
MAP_AREA_H = MAP_AREA_BOTTOM - MAP_AREA_TOP

ISO_ORIGIN_X = MAP_AREA_LEFT + MAP_AREA_W // 2
ISO_ORIGIN_Y = MAP_AREA_TOP + 40


def apply_zoom(new_zoom, pivot_x=None, pivot_y=None):
    global CAM_ZOOM, TILE_W, TILE_H, CAM_PAN_X, CAM_PAN_Y
    new_zoom = max(CAM_ZOOM_MIN, min(CAM_ZOOM_MAX, new_zoom))
    if new_zoom == CAM_ZOOM:
        return
    if pivot_x is not None and pivot_y is not None:
        world_x = (pivot_x - ISO_ORIGIN_X - CAM_PAN_X)
        world_y = (pivot_y - ISO_ORIGIN_Y - CAM_PAN_Y)
        scale = new_zoom / CAM_ZOOM
        CAM_PAN_X = int(pivot_x - ISO_ORIGIN_X - world_x * scale)
        CAM_PAN_Y = int(pivot_y - ISO_ORIGIN_Y - world_y * scale)
    CAM_ZOOM = new_zoom
    TILE_W = max(8, int(TILE_W_BASE * CAM_ZOOM))
    TILE_H = max(4, int(TILE_H_BASE * CAM_ZOOM))


def camera_reset():
    global CAM_ZOOM, CAM_PAN_X, CAM_PAN_Y, TILE_W, TILE_H
    CAM_ZOOM = 1.0
    CAM_PAN_X = 0
    CAM_PAN_Y = 0
    TILE_W = TILE_W_BASE
    TILE_H = TILE_H_BASE


# Состояния клеток
CELL_EMPTY    = 0
CELL_FOREST   = 1
CELL_WATER    = 2
CELL_VILLAGE  = 3
CELL_FIREBREAK = 4
CELL_BURNING  = 5
CELL_BURNT    = 6
CELL_ROAD     = 7

# Типы леса
F_NONE   = 0
F_PINE   = 1
F_SPRUCE = 2
F_BIRCH  = 3
F_PEAT   = 4

# Цвета
COL_BG        = (14, 22, 26)
COL_BG_LIGHT  = (22, 32, 38)
COL_PANEL     = (26, 36, 42)
COL_PANEL_HI  = (36, 48, 56)
COL_TEXT      = (220, 230, 235)
COL_TEXT_DIM  = (140, 155, 165)
COL_ACCENT    = (255, 170, 50)
COL_ACCENT2   = (80, 170, 240)
COL_DANGER    = (235, 70, 60)
COL_OK        = (100, 200, 120)
COL_WARN      = (245, 200, 80)

COL_MEADOW      = (140, 170, 100)
COL_MEADOW_DK   = (110, 140, 75)
COL_PINE        = (55, 95, 65)
COL_PINE_DK     = (35, 70, 45)
COL_SPRUCE      = (40, 75, 55)
COL_SPRUCE_DK   = (25, 55, 40)
COL_BIRCH       = (150, 175, 100)
COL_BIRCH_DK    = (120, 145, 75)
COL_PEAT        = (85, 65, 45)
COL_PEAT_DK     = (60, 45, 30)
COL_WATER       = (70, 110, 160)
COL_WATER_DK    = (45, 80, 125)
COL_VILLAGE     = (180, 140, 110)
COL_VILLAGE_RF  = (150, 60, 55)
COL_FIREBREAK   = (165, 140, 105)
COL_FIREBREAK_D = (125, 105, 75)
COL_BURNING     = (240, 110, 40)
COL_BURNING_HOT = (255, 215, 80)
COL_BURNT       = (40, 35, 32)
COL_BURNT_DK    = (25, 22, 20)
COL_ROAD        = (95, 85, 70)

FOREST_DATA = {
    F_PINE: {
        "name": "Сосняк", "class": "I", "p_base": 0.55,
        "burn_time": 3, "biomass": 180,
        "color": COL_PINE, "color_dk": COL_PINE_DK, "tree_h": 8,
    },
    F_SPRUCE: {
        "name": "Ельник", "class": "II", "p_base": 0.42,
        "burn_time": 4, "biomass": 220,
        "color": COL_SPRUCE, "color_dk": COL_SPRUCE_DK, "tree_h": 10,
    },
    F_BIRCH: {
        "name": "Березняк", "class": "III", "p_base": 0.22,
        "burn_time": 3, "biomass": 140,
        "color": COL_BIRCH, "color_dk": COL_BIRCH_DK, "tree_h": 7,
    },
    F_PEAT: {
        "name": "Торфяник", "class": "особый", "p_base": 0.12,
        "burn_time": 12, "biomass": 95,
        "color": COL_PEAT, "color_dk": COL_PEAT_DK, "tree_h": 2,
    },
}

TOOL_WATCHTOWER  = "watchtower"
TOOL_SENSOR      = "sensor"
TOOL_FIREBREAK   = "firebreak"
TOOL_BRIGADE     = "brigade"
TOOL_PLANE       = "plane"

TOOLS = {
    TOOL_WATCHTOWER: {
        "name": "Вышка наблюдения", "cost": 300_000, "radius": 8,
        "desc": "Радиус 8 клеток. Обнаруживает пожары днём.",
    },
    TOOL_SENSOR: {
        "name": "Датчик дыма", "cost": 50_000, "radius": 3,
        "desc": "Радиус 3 клетки. Круглосуточно, срабатывает мгновенно.",
    },
    TOOL_FIREBREAK: {
        "name": "Минер. полоса (1 км)", "cost": 100_000,
        "desc": "Снижает вероятность перехода огня до 0.05.",
    },
    TOOL_BRIGADE: {
        "name": "Наземная бригада", "cost": 2_000_000, "move_range": 4,
        "desc": "Тушит 1 клетку за ход или прокладывает полосу.",
    },
    TOOL_PLANE: {
        "name": "Бе-200 (база)", "cost": 15_000_000,
        "drop_cost": 500_000, "radius": 3,
        "desc": "База: 15 млн ₽. Сброс: 500 тыс ₽ за 3 клетки. Нужен водоём.",
    },
}

STARTING_BUDGET = 50_000_000


def grid_to_screen(gx, gy):
    sx = ISO_ORIGIN_X + CAM_PAN_X + (gx - gy) * TILE_W // 2
    sy = ISO_ORIGIN_Y + CAM_PAN_Y + (gx + gy) * TILE_H // 2
    return sx, sy

def screen_to_grid(sx, sy):
    dx = sx - ISO_ORIGIN_X - CAM_PAN_X
    dy = sy - ISO_ORIGIN_Y - CAM_PAN_Y
    gx = (dx / (TILE_W / 2) + dy / (TILE_H / 2)) / 2
    gy = (dy / (TILE_H / 2) - dx / (TILE_W / 2)) / 2
    gx_i, gy_i = int(round(gx)), int(round(gy))
    if 0 <= gx_i < GRID_W and 0 <= gy_i < GRID_H:
        return gx_i, gy_i
    return None

def tile_polygon(sx, sy):
    return [
        (sx, sy - TILE_H // 2),
        (sx + TILE_W // 2, sy),
        (sx, sy + TILE_H // 2),
        (sx - TILE_W // 2, sy),
    ]


# ═══════════════════════════════════════════════════════════════════
#   КЛЕТКА (с фиксом forest_original)
# ═══════════════════════════════════════════════════════════════════

class Cell:
    __slots__ = ("state", "forest", "forest_original", "burn_time",
                 "wet_turns", "visible", "height", "humidity_local",
                 "last_p")  # last_p — отладочная вероятность для слоу-мо

    def __init__(self):
        self.state = CELL_EMPTY
        self.forest = F_NONE
        self.forest_original = F_NONE  # ФИКС: сохраняем тип леса до горения
        self.burn_time = 0
        self.wet_turns = 0
        self.visible = False
        self.height = 0
        self.humidity_local = 0
        self.last_p = 0.0  # для отладочной визуализации


# ═══════════════════════════════════════════════════════════════════
#   ПОГОДА
# ═══════════════════════════════════════════════════════════════════

WIND_DIRS = {
    "N":  (0, -1), "NE": (1, -1), "E":  (1, 0), "SE": (1, 1),
    "S":  (0, 1),  "SW": (-1, 1), "W":  (-1, 0), "NW": (-1, -1),
}
WIND_DIR_NAMES = list(WIND_DIRS.keys())
WIND_ARROWS = {
    "N": "↑", "NE": "↗", "E": "→", "SE": "↘",
    "S": "↓", "SW": "↙", "W": "←", "NW": "↖",
}


class Weather:
    def __init__(self, seed=None):
        self.rng = random.Random(seed)
        self.wind_dir = "SW"
        self.wind_speed = 5
        self.humidity = 55
        self.temp = 22
        self.nesterov = 0.0
        self.rain_today = 0.0
        self._streak_dry = 0

    def advance(self, day):
        season_t = math.sin(day / 90 * math.pi)
        base_temp = 18 + 12 * season_t
        self.temp = max(10, base_temp + self.rng.uniform(-4, 4))

        if self.rng.random() < 0.14:
            self.rain_today = self.rng.uniform(1, 20)
        else:
            self.rain_today = 0

        hum_base = 80 - self.temp * 1.2
        if self.rain_today > 3:
            hum_base += 20
        self.humidity = max(15, min(95, hum_base + self.rng.uniform(-8, 8)))

        if self.rng.random() < 0.25:
            idx = WIND_DIR_NAMES.index(self.wind_dir)
            shift = self.rng.choice([-1, 1])
            self.wind_dir = WIND_DIR_NAMES[(idx + shift) % 8]
        self.wind_speed = max(1, min(25,
            self.wind_speed + self.rng.uniform(-3, 3)))

        if self.rain_today > 3:
            self.nesterov = 0
            self._streak_dry = 0
        else:
            t_dew = self.temp - (100 - self.humidity) / 5
            self.nesterov += self.temp * (self.temp - t_dew)
            self._streak_dry += 1

    def fire_danger_class(self):
        n = self.nesterov
        if n < 300:   return "I",   "низкая"
        if n < 1000:  return "II",  "малая"
        if n < 4000:  return "III", "средняя"
        if n < 10000: return "IV",  "высокая"
        return "V", "чрезвычайная"

    def lightning_prob_per_cell(self):
        base = 4e-5
        if self.nesterov > 500:
            base *= (1 + self.nesterov / 1500)
        if self.rain_today > 0:
            base *= 4
        return min(base, 8e-4)

    def spread_multiplier(self, dx_wind, dy_wind, cell_dx, cell_dy):
        dot = dx_wind * cell_dx + dy_wind * cell_dy
        wind_factor = 1.0 + 0.08 * self.wind_speed * dot / math.sqrt(max(cell_dx**2 + cell_dy**2, 1))
        wind_factor = max(0.1, wind_factor)

        if self.humidity < 30:
            hum_factor = 1.6
        elif self.humidity > 70:
            hum_factor = 0.4
        else:
            hum_factor = 1.0 - (self.humidity - 30) / 80

        return wind_factor * hum_factor


# ═══════════════════════════════════════════════════════════════════
#   СЕТКА И АВТОМАТ
# ═══════════════════════════════════════════════════════════════════

class Grid:
    def __init__(self, seed=None):
        self.w, self.h = GRID_W, GRID_H
        self.cells = [[Cell() for _ in range(self.h)] for _ in range(self.w)]
        self.rng = random.Random(seed)
        self.initial_forest_ha = {F_PINE: 0, F_SPRUCE: 0, F_BIRCH: 0, F_PEAT: 0}
        self.burnt_by_type = {F_PINE: 0, F_SPRUCE: 0, F_BIRCH: 0, F_PEAT: 0}
        self.total_ha_burnt = 0
        self.village_cells = []

    def cell(self, x, y):
        return self.cells[x][y]

    def in_bounds(self, x, y):
        return 0 <= x < self.w and 0 <= y < self.h

    def neighbors4(self, x, y):
        for dx, dy in ((0, -1), (1, 0), (0, 1), (-1, 0)):
            nx, ny = x + dx, y + dy
            if self.in_bounds(nx, ny):
                yield nx, ny, dx, dy

    def neighbors8(self, x, y):
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                if dx == 0 and dy == 0:
                    continue
                nx, ny = x + dx, y + dy
                if self.in_bounds(nx, ny):
                    yield nx, ny, dx, dy

    def ignite(self, x, y):
        c = self.cell(x, y)
        if c.state == CELL_FOREST:
            c.state = CELL_BURNING
            c.burn_time = FOREST_DATA[c.forest]["burn_time"]
            c.forest_original = c.forest  # сохраняем для возможного тушения
            return True
        return False

    def step_fire(self, weather):
        """Один шаг клеточного автомата пожара (с двойной буферизацией)."""
        new_ignitions = []
        to_burnout = []

        wdx, wdy = WIND_DIRS[weather.wind_dir]

        for x in range(self.w):
            for y in range(self.h):
                c = self.cells[x][y]
                if c.state == CELL_BURNING:
                    c.burn_time -= 1
                    if c.burn_time <= 0:
                        to_burnout.append((x, y))
                    else:
                        for nx, ny, dx, dy in self.neighbors4(x, y):
                            nc = self.cells[nx][ny]
                            if nc.state != CELL_FOREST:
                                continue
                            if nc.wet_turns > 0:
                                nc.last_p = 0.0
                                continue
                            p = FOREST_DATA[nc.forest]["p_base"]
                            p *= weather.spread_multiplier(wdx, wdy, dx, dy)
                            p = min(p, 1.0)
                            nc.last_p = max(nc.last_p, p)  # для визуализации
                            if self.rng.random() < p:
                                new_ignitions.append((nx, ny))

        for x, y in to_burnout:
            c = self.cells[x][y]
            c.state = CELL_BURNT
            self.total_ha_burnt += 1
            if c.forest != F_NONE:
                self.burnt_by_type[c.forest] = self.burnt_by_type.get(c.forest, 0) + 1
            c.forest = F_NONE
            c.forest_original = F_NONE

        for x, y in new_ignitions:
            c = self.cells[x][y]
            if c.state == CELL_FOREST:
                c.state = CELL_BURNING
                c.burn_time = FOREST_DATA[c.forest]["burn_time"]
                c.forest_original = c.forest  # ФИКС: запоминаем тип

        for x in range(self.w):
            for y in range(self.h):
                c = self.cells[x][y]
                if c.wet_turns > 0:
                    c.wet_turns -= 1

    def village_burning(self):
        for x, y in self.village_cells:
            c = self.cells[x][y]
            if c.state == CELL_BURNING or c.state == CELL_BURNT:
                return True
        return False

    def extinguish(self, x, y):
        """Тушение клетки. ФИКС: используем forest_original."""
        c = self.cell(x, y)
        if c.state == CELL_BURNING:
            c.state = CELL_FOREST
            c.burn_time = 0
            c.wet_turns = 3
            # forest сохранён, forest_original тоже
            return True
        return False

    def count_visible(self):
        return sum(1 for x in range(self.w) for y in range(self.h)
                   if self.cells[x][y].visible)

    def remaining_forest_ha(self):
        return sum(1 for x in range(self.w) for y in range(self.h)
                   if self.cells[x][y].state == CELL_FOREST)

    def forest_density(self):
        total = self.w * self.h
        return self.remaining_forest_ha() / total

    def shannon_index(self):
        counts = {F_PINE: 0, F_SPRUCE: 0, F_BIRCH: 0, F_PEAT: 0}
        total = 0
        for x in range(self.w):
            for y in range(self.h):
                c = self.cells[x][y]
                if c.state == CELL_FOREST and c.forest in counts:
                    counts[c.forest] += 1
                    total += 1
        if total == 0:
            return 0
        h = 0
        for n in counts.values():
            if n > 0:
                p = n / total
                h -= p * math.log(p)
        return h


# ═══════════════════════════════════════════════════════════════════
#   ГЕНЕРАЦИЯ КАРТЫ — ПАРАМЕТРИЗОВАНА
# ═══════════════════════════════════════════════════════════════════

# Типы карт для разных экспериментов
MAP_TYPES = {
    "balanced": {
        "name": "Сбалансированная",
        "desc": "Все типы леса, посёлок, водоём",
        "has_village": True,
        "has_water": True,
        "has_road": True,
        "type_mix": {F_PINE: 0.6, F_SPRUCE: 0.2, F_BIRCH: 0.15, F_PEAT: 0.05},
    },
    "monoculture": {
        "name": "Монокультура (сосняк)",
        "desc": "Только сосняк — для чистого опыта по перколяции",
        "has_village": True,
        "has_water": True,
        "has_road": True,
        "type_mix": {F_PINE: 1.0, F_SPRUCE: 0.0, F_BIRCH: 0.0, F_PEAT: 0.0},
    },
    "mixed": {
        "name": "Смешанный лес",
        "desc": "Лиственные превалируют — низкий p_base",
        "has_village": True,
        "has_water": True,
        "has_road": True,
        "type_mix": {F_PINE: 0.25, F_SPRUCE: 0.15, F_BIRCH: 0.55, F_PEAT: 0.05},
    },
    "lab": {
        "name": "Лаборатория",
        "desc": "Чистая решётка для опытов: ни посёлка, ни дорог",
        "has_village": False,
        "has_water": False,
        "has_road": False,
        "type_mix": {F_PINE: 1.0, F_SPRUCE: 0.0, F_BIRCH: 0.0, F_PEAT: 0.0},
    },
}


def generate_map(grid, seed=42, density=0.7, map_type="balanced"):
    """
    Параметризованная генерация:
      density   — целевая доля клеток с лесом (0.1..0.95)
      map_type  — ключ из MAP_TYPES
    """
    rng = random.Random(seed)
    grid.rng = random.Random(seed)
    mt = MAP_TYPES[map_type]

    # Сброс
    for x in range(grid.w):
        for y in range(grid.h):
            c = grid.cells[x][y]
            c.state = CELL_EMPTY
            c.forest = F_NONE
            c.forest_original = F_NONE
    grid.village_cells = []
    grid.initial_forest_ha = {F_PINE: 0, F_SPRUCE: 0, F_BIRCH: 0, F_PEAT: 0}
    grid.burnt_by_type = {F_PINE: 0, F_SPRUCE: 0, F_BIRCH: 0, F_PEAT: 0}
    grid.total_ha_burnt = 0

    # Водоём
    if mt["has_water"]:
        lake_cx, lake_cy = 10, 30
        for x in range(grid.w):
            for y in range(grid.h):
                d2 = (x - lake_cx) ** 2 + (y - lake_cy) ** 2
                if d2 < 25 + rng.uniform(-4, 4):
                    grid.cells[x][y].state = CELL_WATER
        # Река
        rx, ry = lake_cx + 5, lake_cy
        for _ in range(80):
            if not grid.in_bounds(rx, ry):
                break
            grid.cells[rx][ry].state = CELL_WATER
            if rng.random() < 0.7:
                rx += 1
            else:
                ry += rng.choice([-1, 1])
            ry = max(0, min(grid.h - 1, ry))

    # Посёлок
    if mt["has_village"]:
        vil_x, vil_y = 30, 20
        for dx in range(-2, 3):
            for dy in range(-1, 2):
                x, y = vil_x + dx, vil_y + dy
                if grid.in_bounds(x, y):
                    grid.cells[x][y].state = CELL_VILLAGE
                    grid.village_cells.append((x, y))

        # Дорога
        if mt["has_road"]:
            rx, ry = vil_x + 3, vil_y
            while grid.in_bounds(rx, ry):
                if grid.cells[rx][ry].state == CELL_EMPTY:
                    grid.cells[rx][ry].state = CELL_ROAD
                rx += 1

    # Расставляем лес по type_mix и density
    type_mix = mt["type_mix"]
    type_keys = [k for k, v in type_mix.items() if v > 0]
    type_weights = [type_mix[k] for k in type_keys]

    for x in range(grid.w):
        for y in range(grid.h):
            c = grid.cells[x][y]
            if c.state != CELL_EMPTY:
                continue
            if rng.random() < density:
                ftype = rng.choices(type_keys, weights=type_weights, k=1)[0]
                c.state = CELL_FOREST
                c.forest = ftype
                c.forest_original = ftype
                grid.initial_forest_ha[ftype] = grid.initial_forest_ha.get(ftype, 0) + 1

    return grid


# Старая функция-обёртка для совместимости
def generate_tutorial_map(grid, seed=42):
    return generate_map(grid, seed=seed, density=0.78, map_type="balanced")


# ═══════════════════════════════════════════════════════════════════
#   ИНСТРУМЕНТЫ — ОБЪЕКТЫ НА КАРТЕ
# ═══════════════════════════════════════════════════════════════════

class Placement:
    def __init__(self, kind, x, y, extra=None):
        self.kind = kind
        self.x = x
        self.y = y
        self.extra = extra or {}


# ═══════════════════════════════════════════════════════════════════
#   НАСТРОЙКИ НОВОЙ ИГРЫ
# ═══════════════════════════════════════════════════════════════════

class GameSetup:
    """Параметры, которые игрок выбирает перед началом."""
    def __init__(self):
        self.seed = 42
        self.density = 0.78
        self.map_type = "balanced"
        self.budget = STARTING_BUDGET
        self.max_days = 60


# ═══════════════════════════════════════════════════════════════════
#   ЗАПИСЬ ЭКСПЕРИМЕНТА (для лаборатории)
# ═══════════════════════════════════════════════════════════════════

class Experiment:
    """Один эксперимент в лаборатории — начальные условия и результат."""
    def __init__(self, density, map_type, wind_speed, humidity, seed,
                 burnt_frac, total_burnt, total_initial, steps):
        self.density = density
        self.map_type = map_type
        self.wind_speed = wind_speed
        self.humidity = humidity
        self.seed = seed
        self.burnt_frac = burnt_frac
        self.total_burnt = total_burnt
        self.total_initial = total_initial
        self.steps = steps
        self.timestamp = datetime.now().strftime("%H:%M:%S")


# ═══════════════════════════════════════════════════════════════════
#   ИГРОВОЕ СОСТОЯНИЕ
# ═══════════════════════════════════════════════════════════════════

class GameState:
    def __init__(self, setup=None):
        if setup is None:
            setup = GameSetup()
        self.setup = setup
        self.seed = setup.seed
        self.grid = Grid(setup.seed)
        generate_map(self.grid, seed=setup.seed,
                     density=setup.density, map_type=setup.map_type)
        self.weather = Weather(setup.seed)
        self.day = 0
        self.max_days = setup.max_days
        self.budget = setup.budget
        self.initial_budget = setup.budget
        self.placements = []
        self.active_fires = 0
        self.history = []
        self.events_log = deque(maxlen=15)
        self.phase = "menu"
        self.selected_tool = None
        self.firebreak_anchor = None
        self.mouse_grid = None
        self.selected_unit = None
        self.plane_step = None
        self.end_reason = None
        self.brigade_moves_left = {}

        # НОВОЕ: режимы отладки
        self.slow_mo = False           # пошаговый показ переходов
        self.show_probabilities = False  # подсветка p на соседях огня

        # НОВОЕ: встроенный мини-эксперимент
        self.inline_experiments = []   # точки на встроенном графике

        self._log("Добро пожаловать в Сосновский лесхоз.")
        self._log("Фаза подготовки: разместите средства защиты.")

    def _log(self, msg):
        self.events_log.append(f"Д.{self.day}: {msg}")

    def can_place(self, tool, x, y):
        cost = TOOLS[tool]["cost"]
        if self.budget < cost:
            return False, "Недостаточно бюджета"
        if not self.grid.in_bounds(x, y):
            return False, "Вне карты"
        c = self.grid.cell(x, y)
        if tool == TOOL_WATCHTOWER or tool == TOOL_SENSOR:
            if c.state in (CELL_WATER, CELL_BURNING, CELL_BURNT, CELL_VILLAGE):
                return False, "Нельзя размещать здесь"
        if tool == TOOL_FIREBREAK:
            if c.state not in (CELL_FOREST, CELL_EMPTY):
                return False, "Полоса прокладывается по лесу или лугу"
        if tool == TOOL_BRIGADE:
            if c.state in (CELL_WATER, CELL_BURNING):
                return False, "Бригада не может быть на воде или в огне"
        if tool == TOOL_PLANE:
            has_water = False
            for nx, ny, _, _ in self.grid.neighbors8(x, y):
                if self.grid.cell(nx, ny).state == CELL_WATER:
                    has_water = True
                    break
            if not has_water:
                return False, "База Бе-200 требует водоём рядом"
            if c.state != CELL_EMPTY:
                return False, "Место должно быть свободно"
        return True, ""

    def place(self, tool, x, y):
        ok, msg = self.can_place(tool, x, y)
        if not ok:
            self._log(msg)
            return False
        cost = TOOLS[tool]["cost"]
        self.budget -= cost
        p = Placement(tool, x, y)
        self.placements.append(p)

        if tool == TOOL_FIREBREAK:
            self.grid.cell(x, y).state = CELL_FIREBREAK
            self.grid.cell(x, y).forest = F_NONE
            self.grid.cell(x, y).forest_original = F_NONE

        self._recalc_visibility()
        self._log(f"Установлено: {TOOLS[tool]['name']}")
        return True

    def _recalc_visibility(self):
        for x in range(self.grid.w):
            for y in range(self.grid.h):
                self.grid.cells[x][y].visible = False
        for p in self.placements:
            r = 0
            if p.kind == TOOL_WATCHTOWER:
                r = TOOLS[TOOL_WATCHTOWER]["radius"]
            elif p.kind == TOOL_SENSOR:
                r = TOOLS[TOOL_SENSOR]["radius"]
            if r > 0:
                for x in range(max(0, p.x - r), min(self.grid.w, p.x + r + 1)):
                    for y in range(max(0, p.y - r), min(self.grid.h, p.y + r + 1)):
                        if (x - p.x) ** 2 + (y - p.y) ** 2 <= r * r:
                            self.grid.cells[x][y].visible = True

    def advance_day(self):
        if self.phase != "play":
            return
        self.day += 1
        self.weather.advance(self.day)
        # Сброс last_p перед новым шагом
        for x in range(self.grid.w):
            for y in range(self.grid.h):
                self.grid.cells[x][y].last_p = 0.0

        lp = self.weather.lightning_prob_per_cell()
        for x in range(self.grid.w):
            for y in range(self.grid.h):
                c = self.grid.cell(x, y)
                if c.state == CELL_FOREST and self.grid.rng.random() < lp:
                    c.state = CELL_BURNING
                    c.burn_time = FOREST_DATA[c.forest]["burn_time"]
                    c.forest_original = c.forest
                    self._log(f"⚡ Удар молнии в сектор ({x},{y})")

        if self.grid.rng.random() < 0.12 and self.weather.nesterov > 200:
            candidates = []
            for x in range(self.grid.w):
                for y in range(self.grid.h):
                    c = self.grid.cell(x, y)
                    if c.state == CELL_FOREST:
                        for nx, ny, _, _ in self.grid.neighbors8(x, y):
                            if self.grid.cell(nx, ny).state == CELL_ROAD:
                                candidates.append((x, y))
                                break
            if candidates:
                x, y = self.grid.rng.choice(candidates)
                c = self.grid.cell(x, y)
                c.state = CELL_BURNING
                c.burn_time = FOREST_DATA[c.forest]["burn_time"]
                c.forest_original = c.forest
                self._log(f"🔥 Неосторожное обращение с огнём у дороги ({x},{y})")

        self.grid.step_fire(self.weather)

        self.active_fires = sum(1 for x in range(self.grid.w)
                                for y in range(self.grid.h)
                                if self.grid.cell(x, y).state == CELL_BURNING)

        self.brigade_moves_left = {}
        for i, p in enumerate(self.placements):
            if p.kind == TOOL_BRIGADE:
                self.brigade_moves_left[i] = TOOLS[TOOL_BRIGADE]["move_range"]

        self.history.append({
            "day": self.day,
            "temp": round(self.weather.temp, 1),
            "humidity": round(self.weather.humidity, 1),
            "wind_dir": self.weather.wind_dir,
            "wind_speed": round(self.weather.wind_speed, 1),
            "rain": round(self.weather.rain_today, 1),
            "nesterov": round(self.weather.nesterov, 0),
            "active_fires": self.active_fires,
            "burnt_total": self.grid.total_ha_burnt,
            "forest_remaining": self.grid.remaining_forest_ha(),
            "budget": self.budget,
        })

        if self.grid.village_burning():
            self.end_reason = "village"
            self.phase = "analysis"
            self._log("❌ Посёлок охвачен огнём! Миссия провалена.")
        elif self.day >= self.max_days:
            self.end_reason = "days"
            self.phase = "analysis"
            self._log("✅ Сезон завершён.")

    def brigade_act(self, brigade_idx, target_x, target_y):
        if brigade_idx not in self.brigade_moves_left:
            return False
        if self.brigade_moves_left[brigade_idx] <= 0:
            self._log("У бригады закончились ходы на сегодня")
            return False
        p = self.placements[brigade_idx]
        dist = abs(p.x - target_x) + abs(p.y - target_y)
        if dist > self.brigade_moves_left[brigade_idx]:
            self._log(f"Слишком далеко (дистанция {dist})")
            return False
        self.brigade_moves_left[brigade_idx] -= dist
        p.x, p.y = target_x, target_y
        extinguished = 0
        for nx, ny, _, _ in self.grid.neighbors8(p.x, p.y):
            if self.grid.cell(nx, ny).state == CELL_BURNING:
                fc = self.grid.cell(nx, ny)
                # ФИКС: используем forest_original, чтобы не потерять тип
                if fc.forest_original != F_NONE and fc.forest == F_NONE:
                    fc.forest = fc.forest_original
                fc.state = CELL_FOREST
                fc.burn_time = 0
                fc.wet_turns = 2
                extinguished += 1
                if extinguished >= 1:
                    break
        if extinguished:
            self._log(f"Бригада потушила очаг в ({p.x},{p.y})")
        return True

    def plane_drop(self, base_idx, target_x, target_y):
        drop_cost = TOOLS[TOOL_PLANE]["drop_cost"]
        if self.budget < drop_cost:
            self._log("Недостаточно бюджета для вылета")
            return False
        self.budget -= drop_cost
        r = TOOLS[TOOL_PLANE]["radius"]
        cnt = 0
        for dx in range(-r, r + 1):
            for dy in range(-r, r + 1):
                if dx*dx + dy*dy > r*r:
                    continue
                x, y = target_x + dx, target_y + dy
                if not self.grid.in_bounds(x, y):
                    continue
                c = self.grid.cell(x, y)
                if c.state == CELL_BURNING:
                    if c.forest_original != F_NONE and c.forest == F_NONE:
                        c.forest = c.forest_original
                    c.state = CELL_FOREST
                    c.burn_time = 0
                    c.wet_turns = 4
                    cnt += 1
                elif c.state == CELL_FOREST:
                    c.wet_turns = 2
        self._log(f"Бе-200 сбросил воду, потушено {cnt} клеток")
        return True

    def start_season(self):
        if self.phase == "build":
            self.phase = "play"
            self._recalc_visibility()
            self._log("Начало пожароопасного сезона. Удачи!")
            candidates = []
            for x in range(self.grid.w):
                for y in range(self.grid.h):
                    c = self.grid.cell(x, y)
                    if c.state == CELL_FOREST and c.forest == F_PINE:
                        if self.grid.village_cells:
                            min_dist = min(abs(x - vx) + abs(y - vy)
                                           for vx, vy in self.grid.village_cells)
                            if 10 <= min_dist <= 20:
                                candidates.append((x, y))
                        else:
                            candidates.append((x, y))
            if candidates:
                x, y = self.grid.rng.choice(candidates)
                c = self.grid.cell(x, y)
                c.state = CELL_BURNING
                c.burn_time = FOREST_DATA[c.forest]["burn_time"]
                c.forest_original = c.forest
                self._log(f"⚠ Обнаружен очаг в секторе ({x},{y})!")

    def compute_score(self):
        g = self.grid
        total_forest = sum(g.initial_forest_ha.values())
        saved = total_forest - g.total_ha_burnt
        saved = max(0, saved)
        saved_frac = saved / total_forest if total_forest > 0 else 0
        shannon = g.shannon_index()
        spent = self.initial_budget - self.budget
        score = (saved * shannon) / (spent / 1_000_000 + 1)
        if self.end_reason == "village":
            score *= 0.3
        tier = "—"
        if self.end_reason != "village":
            if saved_frac > 0.9 and shannon > 1.0:
                tier = "GOLD"
            elif saved_frac > 0.7:
                tier = "SILVER"
            elif saved_frac > 0.5:
                tier = "BRONZE"
        return score, saved_frac, shannon, tier


# ═══════════════════════════════════════════════════════════════════
#   ЛАБОРАТОРИЯ ПЕРКОЛЯЦИИ
# ═══════════════════════════════════════════════════════════════════

class PercolationLab:
    """
    Отдельный режим — чистый эксперимент.
    
    Школьник:
      1. Задаёт параметры (плотность, ветер, влажность, тип леса)
      2. Запускает 1 эксперимент или серию из N
      3. Видит точки на графике "плотность vs % сгоревшего"
      4. Может включить визуализацию или прогон без графики
    """

    def __init__(self):
        # Параметры опыта (изменяются ползунками)
        self.density = 0.5
        self.wind_speed = 0  # 0 = без ветра, чистая перколяция
        self.humidity = 50
        self.forest_type = F_PINE  # для чистоты опыта — монокультура
        self.seed = 42

        # Результаты
        self.experiments = []  # list[Experiment]

        # Состояние симуляции
        self.grid = None
        self.weather = None
        self.is_running = False
        self.current_step = 0
        self.steps_burnt = 0
        self.steps_initial = 0

        # Автопрогон
        self.batch_remaining = 0
        self.batch_animate = False
        self.batch_total = 0
        self.batch_target_density = None  # None = случайная плотность; число = фикс
        self.last_step_time = 0
        self.step_delay_ms = 80  # задержка между шагами при анимации

        # UI
        self.show_graph_mode = "scatter"  # scatter / table

    def setup_experiment(self, density=None, animate=True, seed=None):
        """Подготавливает решётку под эксперимент."""
        if density is None:
            density = self.density
        if seed is None:
            seed = random.randint(0, 100000)
        self.seed = seed
        rng = random.Random(seed)

        self.grid = Grid(seed)
        # Чистая лабораторная карта — только лес одного типа
        for x in range(GRID_W):
            for y in range(GRID_H):
                c = self.grid.cells[x][y]
                c.state = CELL_EMPTY
                c.forest = F_NONE
                c.forest_original = F_NONE
                if rng.random() < density:
                    c.state = CELL_FOREST
                    c.forest = self.forest_type
                    c.forest_original = self.forest_type
                    self.grid.initial_forest_ha[self.forest_type] = \
                        self.grid.initial_forest_ha.get(self.forest_type, 0) + 1

        # Поджигаем центр
        cx, cy = GRID_W // 2, GRID_H // 2
        if self.grid.cells[cx][cy].state == CELL_FOREST:
            self.grid.ignite(cx, cy)
        else:
            # Ищем ближайшую лесную клетку
            for r in range(1, max(GRID_W, GRID_H)):
                found = False
                for dx in range(-r, r + 1):
                    for dy in range(-r, r + 1):
                        x, y = cx + dx, cy + dy
                        if self.grid.in_bounds(x, y) and self.grid.cells[x][y].state == CELL_FOREST:
                            self.grid.ignite(x, y)
                            found = True
                            break
                    if found:
                        break
                if found:
                    break

        # Погода — статичная для воспроизводимости
        self.weather = Weather(seed)
        self.weather.wind_speed = self.wind_speed
        self.weather.humidity = self.humidity
        self.weather.wind_dir = "N"  # нейтральное

        self.is_running = True
        self.current_step = 0
        self.steps_initial = self.grid.remaining_forest_ha()
        self.steps_burnt = 0
        self.last_step_time = 0
        self.batch_animate = animate
        self.batch_target_density = density

    def step(self):
        """Один шаг автомата. Возвращает True, если симуляция продолжается."""
        if not self.is_running or self.grid is None:
            return False

        active = sum(1 for x in range(GRID_W) for y in range(GRID_H)
                     if self.grid.cells[x][y].state == CELL_BURNING)
        if active == 0:
            self._finalize_experiment()
            return False

        self.grid.step_fire(self.weather)
        self.current_step += 1

        # Защита от бесконечных циклов (торфяники могут гореть долго)
        if self.current_step > 200:
            self._finalize_experiment()
            return False

        return True

    def run_to_completion(self):
        """Прогоняет эксперимент целиком без анимации (быстро)."""
        if not self.is_running:
            return
        while self.step():
            pass

    def _finalize_experiment(self):
        """Записывает результат."""
        if self.grid is None or self.steps_initial == 0:
            self.is_running = False
            return
        burnt = self.grid.total_ha_burnt
        frac = burnt / self.steps_initial if self.steps_initial > 0 else 0
        density_used = self.batch_target_density if self.batch_target_density is not None else self.density
        exp = Experiment(
            density=density_used,
            map_type="lab",
            wind_speed=self.weather.wind_speed,
            humidity=self.weather.humidity,
            seed=self.seed,
            burnt_frac=frac,
            total_burnt=burnt,
            total_initial=self.steps_initial,
            steps=self.current_step,
        )
        self.experiments.append(exp)
        self.is_running = False

    def start_batch(self, n, animate=False, density_range=None):
        """
        Запускает серию из n экспериментов.
        Если density_range = (lo, hi) — плотность случайная в диапазоне.
        Если None — используется текущая self.density.
        """
        self.batch_remaining = n
        self.batch_animate = animate
        self.batch_total = n
        self.batch_density_range = density_range

    def update_batch(self):
        """Вызывается каждый кадр в цикле main."""
        if self.batch_remaining <= 0:
            return
        if self.is_running:
            if self.batch_animate:
                # Шаги по таймеру
                now = pygame.time.get_ticks()
                if now - self.last_step_time >= self.step_delay_ms:
                    self.step()
                    self.last_step_time = now
            else:
                # Без анимации — прогоняем целиком за один кадр
                self.run_to_completion()
        else:
            self.batch_remaining -= 1
            if self.batch_remaining > 0:
                # Запускаем следующий
                if self.batch_density_range is not None:
                    lo, hi = self.batch_density_range
                    d = random.uniform(lo, hi)
                else:
                    d = self.density
                self.setup_experiment(density=d, animate=self.batch_animate,
                                      seed=random.randint(0, 1_000_000))

    def stop_batch(self):
        self.batch_remaining = 0
        if self.is_running:
            self._finalize_experiment()

    def clear(self):
        self.experiments = []
        self.grid = None
        self.is_running = False
        self.batch_remaining = 0


# ═══════════════════════════════════════════════════════════════════
#   ВСПОМОГАТЕЛЬНОЕ — ПОДСЧЁТ ВЕРОЯТНОСТИ ДЛЯ ВИЗУАЛИЗАЦИИ
# ═══════════════════════════════════════════════════════════════════

def compute_p_for_neighbor(grid, weather, fx, fy, nx, ny):
    """
    Возвращает (p_base, wind_factor, hum_factor, p_total) для перехода
    от горящей клетки (fx,fy) на лесную (nx,ny).
    Используется для визуализации в слоу-мо и для tooltip'а.
    """
    nc = grid.cell(nx, ny)
    if nc.state != CELL_FOREST:
        return None
    if nc.wet_turns > 0:
        return (0, 0, 0, 0)
    p_base = FOREST_DATA[nc.forest]["p_base"]
    wdx, wdy = WIND_DIRS[weather.wind_dir]
    dx, dy = nx - fx, ny - fy
    dot = wdx * dx + wdy * dy
    wind_factor = 1.0 + 0.08 * weather.wind_speed * dot / math.sqrt(max(dx*dx + dy*dy, 1))
    wind_factor = max(0.1, wind_factor)
    if weather.humidity < 30:
        hum_factor = 1.6
    elif weather.humidity > 70:
        hum_factor = 0.4
    else:
        hum_factor = 1.0 - (weather.humidity - 30) / 80
    p_total = min(p_base * wind_factor * hum_factor, 1.0)
    return (p_base, wind_factor, hum_factor, p_total)


# ═══════════════════════════════════════════════════════════════════
#   ЧАСТЬ 2: РЕНДЕР
# ═══════════════════════════════════════════════════════════════════

_font_cache = {}

def font(size, bold=False):
    key = (size, bold)
    if key not in _font_cache:
        candidates = ["segoeui", "arial", "calibri", "verdana", "tahoma", "dejavusans"]
        f = None
        for name in candidates:
            try:
                f = pygame.font.SysFont(name, size, bold=bold)
                if f:
                    break
            except:
                continue
        if f is None:
            f = pygame.font.Font(None, size)
        _font_cache[key] = f
    return _font_cache[key]


def draw_text(surf, text, pos, size=16, color=COL_TEXT, bold=False, center=False, right=False):
    img = font(size, bold).render(text, True, color)
    rect = img.get_rect()
    if center:
        rect.center = pos
    elif right:
        rect.topright = pos
    else:
        rect.topleft = pos
    surf.blit(img, rect)
    return rect


def darken(color, amount=0.7):
    return tuple(max(0, int(c * amount)) for c in color)

def lighten(color, amount=1.3):
    return tuple(min(255, int(c * amount)) for c in color)


def draw_diamond(surf, cx, cy, color, outline=None):
    pts = [
        (cx, cy - TILE_H // 2),
        (cx + TILE_W // 2, cy),
        (cx, cy + TILE_H // 2),
        (cx - TILE_W // 2, cy),
    ]
    pygame.draw.polygon(surf, color, pts)
    if outline:
        pygame.draw.polygon(surf, outline, pts, 1)


def draw_tile_base(surf, cx, cy, color_top, color_side):
    pygame.draw.polygon(surf, color_side, [
        (cx - TILE_W // 2, cy),
        (cx, cy + TILE_H // 2),
        (cx, cy + TILE_H // 2 + 3),
        (cx - TILE_W // 2, cy + 3),
    ])
    pygame.draw.polygon(surf, darken(color_side, 0.8), [
        (cx + TILE_W // 2, cy),
        (cx, cy + TILE_H // 2),
        (cx, cy + TILE_H // 2 + 3),
        (cx + TILE_W // 2, cy + 3),
    ])
    pts = [
        (cx, cy - TILE_H // 2),
        (cx + TILE_W // 2, cy),
        (cx, cy + TILE_H // 2),
        (cx - TILE_W // 2, cy),
    ]
    pygame.draw.polygon(surf, color_top, pts)


def draw_tree(surf, cx, cy, color, color_dk, height):
    top = (cx, cy - TILE_H // 2 - height)
    base_l = (cx - 3, cy - TILE_H // 2 + 1)
    base_r = (cx + 3, cy - TILE_H // 2 + 1)
    pygame.draw.polygon(surf, color, [top, base_l, base_r])
    pygame.draw.polygon(surf, color_dk, [top, base_l, base_r], 1)


def draw_cell(surf, grid, x, y, game, anim_t, show_p=False):
    cx, cy = grid_to_screen(x, y)
    if cx < MAP_AREA_LEFT - TILE_W or cx > MAP_AREA_RIGHT + TILE_W:
        return
    if cy < MAP_AREA_TOP - TILE_H - 30 or cy > MAP_AREA_BOTTOM + TILE_H:
        return

    c = grid.cell(x, y)

    if c.state == CELL_EMPTY:
        draw_tile_base(surf, cx, cy, COL_MEADOW, COL_MEADOW_DK)

    elif c.state == CELL_WATER:
        phase = math.sin(anim_t * 2 + (x + y) * 0.5)
        col = (
            int(COL_WATER[0] + phase * 10),
            int(COL_WATER[1] + phase * 15),
            int(COL_WATER[2] + phase * 20),
        )
        col = tuple(max(0, min(255, v)) for v in col)
        draw_tile_base(surf, cx, cy, col, COL_WATER_DK)

    elif c.state == CELL_ROAD:
        draw_tile_base(surf, cx, cy, COL_ROAD, darken(COL_ROAD, 0.7))

    elif c.state == CELL_VILLAGE:
        draw_tile_base(surf, cx, cy, COL_VILLAGE, darken(COL_VILLAGE, 0.7))
        roof_w = 6
        roof_h = 5
        house_y = cy - TILE_H // 2 - 4
        pygame.draw.rect(surf, COL_VILLAGE,
                         (cx - roof_w // 2, house_y, roof_w, 6))
        pygame.draw.polygon(surf, COL_VILLAGE_RF, [
            (cx - roof_w // 2 - 1, house_y),
            (cx + roof_w // 2 + 1, house_y),
            (cx, house_y - roof_h),
        ])

    elif c.state == CELL_FIREBREAK:
        draw_tile_base(surf, cx, cy, COL_FIREBREAK, COL_FIREBREAK_D)
        pygame.draw.line(surf, darken(COL_FIREBREAK, 0.6),
                         (cx - 5, cy - 1), (cx + 5, cy + 1), 1)

    elif c.state == CELL_FOREST:
        fdata = FOREST_DATA[c.forest]
        draw_tile_base(surf, cx, cy, COL_MEADOW, COL_MEADOW_DK)
        draw_tree(surf, cx, cy, fdata["color"], fdata["color_dk"], fdata["tree_h"])
        if c.wet_turns > 0:
            overlay = pygame.Surface((TILE_W, TILE_H + 3), pygame.SRCALPHA)
            overlay.fill((80, 140, 220, 60))
            surf.blit(overlay, (cx - TILE_W // 2, cy - TILE_H // 2))
        # НОВОЕ: подсветка вероятности перехода огня
        if show_p and c.last_p > 0.05:
            # Оттенок от жёлтого (низкая p) до красного (высокая)
            t = min(c.last_p, 1.0)
            r = 255
            g = int(220 * (1 - t * 0.7))
            b = 0
            alpha = int(40 + 100 * t)
            overlay = pygame.Surface((TILE_W, TILE_H + 3), pygame.SRCALPHA)
            overlay.fill((r, g, b, alpha))
            surf.blit(overlay, (cx - TILE_W // 2, cy - TILE_H // 2))

    elif c.state == CELL_BURNING:
        flicker = (math.sin(anim_t * 8 + x * 0.7 + y * 0.3) + 1) / 2
        base_col = (
            int(COL_BURNING[0] * (0.8 + flicker * 0.2)),
            int(COL_BURNING[1] * (0.7 + flicker * 0.3)),
            int(COL_BURNING[2] * 0.5),
        )
        base_col = tuple(max(0, min(255, v)) for v in base_col)
        draw_tile_base(surf, cx, cy, base_col, darken(base_col, 0.6))
        h = 6 + int(flicker * 4)
        pygame.draw.polygon(surf, COL_BURNING_HOT, [
            (cx, cy - TILE_H // 2 - h),
            (cx - 3, cy - TILE_H // 2 + 1),
            (cx + 3, cy - TILE_H // 2 + 1),
        ])
        pygame.draw.polygon(surf, COL_BURNING, [
            (cx, cy - TILE_H // 2 - h + 2),
            (cx - 2, cy - TILE_H // 2 + 1),
            (cx + 2, cy - TILE_H // 2 + 1),
        ])

    elif c.state == CELL_BURNT:
        draw_tile_base(surf, cx, cy, COL_BURNT, COL_BURNT_DK)
        pygame.draw.line(surf, (70, 65, 60), (cx - 3, cy - 1), (cx - 3, cy - 4), 1)
        pygame.draw.line(surf, (70, 65, 60), (cx + 2, cy - 1), (cx + 2, cy - 3), 1)


def draw_placement(surf, p, game, anim_t, selected=False):
    cx, cy = grid_to_screen(p.x, p.y)

    if p.kind == TOOL_WATCHTOWER:
        pygame.draw.rect(surf, (210, 210, 210), (cx - 2, cy - 22, 4, 18))
        pygame.draw.rect(surf, (200, 60, 60), (cx - 5, cy - 24, 10, 5))
        pygame.draw.line(surf, (150, 150, 150), (cx - 5, cy - 4), (cx, cy - 22), 1)
        pygame.draw.line(surf, (150, 150, 150), (cx + 5, cy - 4), (cx, cy - 22), 1)
        if selected:
            pygame.draw.circle(surf, COL_ACCENT, (cx, cy - 6), 16, 1)

    elif p.kind == TOOL_SENSOR:
        pygame.draw.rect(surf, (180, 180, 180), (cx - 1, cy - 8, 2, 8))
        pygame.draw.circle(surf, (230, 60, 60), (cx, cy - 9), 2)
        if selected:
            pygame.draw.circle(surf, COL_ACCENT, (cx, cy - 6), 8, 1)

    elif p.kind == TOOL_BRIGADE:
        col = (240, 180, 60) if not selected else COL_ACCENT
        pygame.draw.circle(surf, col, (cx, cy - 6), 6)
        pygame.draw.circle(surf, (0, 0, 0), (cx, cy - 6), 6, 1)
        img = font(11, bold=True).render("Б", True, (0, 0, 0))
        rect = img.get_rect(center=(cx, cy - 6))
        surf.blit(img, rect)

    elif p.kind == TOOL_PLANE:
        pygame.draw.rect(surf, (100, 110, 120), (cx - 8, cy - 10, 16, 8))
        pygame.draw.polygon(surf, (140, 150, 160), [
            (cx - 8, cy - 10), (cx + 8, cy - 10), (cx + 6, cy - 14), (cx - 6, cy - 14)
        ])
        pygame.draw.line(surf, (220, 230, 240), (cx - 5, cy - 17), (cx + 5, cy - 17), 2)
        pygame.draw.line(surf, (220, 230, 240), (cx, cy - 19), (cx, cy - 15), 2)
        if selected:
            pygame.draw.rect(surf, COL_ACCENT, (cx - 10, cy - 20, 20, 22), 1)


def draw_radius(surf, cx_grid, cy_grid, radius, color):
    sx, sy = grid_to_screen(cx_grid, cy_grid)
    rw = radius * TILE_W
    rh = radius * TILE_H
    rect = pygame.Rect(sx - rw, sy - rh, 2 * rw, 2 * rh)
    s = pygame.Surface((2 * rw, 2 * rh), pygame.SRCALPHA)
    pygame.draw.ellipse(s, color, (0, 0, 2 * rw, 2 * rh), 1)
    surf.blit(s, rect.topleft)


def draw_map(surf, game, anim_t, hover_cell=None, show_p=False):
    clip = pygame.Rect(MAP_AREA_LEFT, MAP_AREA_TOP, MAP_AREA_W, MAP_AREA_H)
    surf.set_clip(clip)

    for s in range(game.grid.w + game.grid.h):
        for x in range(game.grid.w):
            y = s - x
            if 0 <= y < game.grid.h:
                draw_cell(surf, game.grid, x, y, game, anim_t, show_p=show_p)

    for i, p in enumerate(game.placements):
        sel = (game.selected_unit is not None and game.selected_unit == i)
        draw_placement(surf, p, game, anim_t, selected=sel)

    if hover_cell:
        hx, hy = hover_cell
        cx, cy = grid_to_screen(hx, hy)
        pts = [
            (cx, cy - TILE_H // 2),
            (cx + TILE_W // 2, cy),
            (cx, cy + TILE_H // 2),
            (cx - TILE_W // 2, cy),
        ]
        pygame.draw.polygon(surf, COL_ACCENT, pts, 2)

    surf.set_clip(None)


def draw_lab_map(surf, lab, anim_t, hover_cell=None):
    """Упрощённая отрисовка решётки в лаборатории."""
    if lab.grid is None:
        # Просто пустая область
        return

    clip = pygame.Rect(MAP_AREA_LEFT, MAP_AREA_TOP, MAP_AREA_W, MAP_AREA_H)
    surf.set_clip(clip)

    for s in range(lab.grid.w + lab.grid.h):
        for x in range(lab.grid.w):
            y = s - x
            if 0 <= y < lab.grid.h:
                # Создаём фейковый game для совместимости
                draw_cell(surf, lab.grid, x, y, None, anim_t, show_p=False)

    if hover_cell:
        hx, hy = hover_cell
        cx, cy = grid_to_screen(hx, hy)
        pts = [
            (cx, cy - TILE_H // 2),
            (cx + TILE_W // 2, cy),
            (cx, cy + TILE_H // 2),
            (cx - TILE_W // 2, cy),
        ]
        pygame.draw.polygon(surf, COL_ACCENT, pts, 2)

    surf.set_clip(None)


def format_money(v):
    if v >= 1_000_000:
        return f"{v/1_000_000:.1f} млн ₽"
    if v >= 1_000:
        return f"{v/1_000:.0f} тыс ₽"
    return f"{v} ₽"


def draw_top_hud(surf, game):
    pygame.draw.rect(surf, COL_PANEL, (0, 0, SCREEN_W, MAP_AREA_TOP))
    pygame.draw.line(surf, COL_PANEL_HI, (0, MAP_AREA_TOP), (SCREEN_W, MAP_AREA_TOP), 1)

    draw_text(surf, "ОГНЕННЫЙ ФРОНТ", (20, 8), 14, COL_TEXT_DIM, bold=True)
    draw_text(surf, "Защита тайги — Сосновский лесхоз", (20, 28), 18, COL_TEXT, bold=True)

    col_x = 440
    draw_text(surf, "ДЕНЬ", (col_x, 10), 11, COL_TEXT_DIM, bold=True)
    draw_text(surf, f"{game.day} / {game.max_days}", (col_x, 26), 22, COL_TEXT, bold=True)

    col_x += 140
    draw_text(surf, "БЮДЖЕТ", (col_x, 10), 11, COL_TEXT_DIM, bold=True)
    bcol = COL_OK if game.budget > 5_000_000 else (COL_WARN if game.budget > 0 else COL_DANGER)
    draw_text(surf, format_money(game.budget), (col_x, 26), 22, bcol, bold=True)

    col_x += 200
    draw_text(surf, "СГОРЕЛО", (col_x, 10), 11, COL_TEXT_DIM, bold=True)
    b = game.grid.total_ha_burnt
    total = sum(game.grid.initial_forest_ha.values())
    pct = 100 * b / total if total > 0 else 0
    bcol = COL_OK if pct < 5 else (COL_WARN if pct < 20 else COL_DANGER)
    draw_text(surf, f"{b} га ({pct:.1f}%)", (col_x, 26), 22, bcol, bold=True)

    col_x += 200
    draw_text(surf, "ОЧАГОВ", (col_x, 10), 11, COL_TEXT_DIM, bold=True)
    fcol = COL_OK if game.active_fires == 0 else COL_DANGER
    draw_text(surf, f"{game.active_fires}", (col_x, 26), 22, fcol, bold=True)

    col_x += 160
    if game.day > 0:
        draw_text(surf, "ИНДЕКС НЕСТЕРОВА", (col_x, 10), 11, COL_TEXT_DIM, bold=True)
        cls, txt = game.weather.fire_danger_class()
        ncol = {"I": COL_OK, "II": COL_OK, "III": COL_WARN, "IV": COL_DANGER, "V": COL_DANGER}[cls]
        draw_text(surf, f"{int(game.weather.nesterov)} ({cls})", (col_x, 26), 22, ncol, bold=True)


def draw_bottom_panel(surf, game):
    pygame.draw.rect(surf, COL_PANEL, (0, MAP_AREA_BOTTOM, SCREEN_W, SCREEN_H - MAP_AREA_BOTTOM))
    pygame.draw.line(surf, COL_PANEL_HI, (0, MAP_AREA_BOTTOM), (SCREEN_W, MAP_AREA_BOTTOM), 1)

    y0 = MAP_AREA_BOTTOM + 14
    draw_text(surf, "ПОГОДА", (20, y0), 12, COL_TEXT_DIM, bold=True)
    w = game.weather
    arrow = WIND_ARROWS[w.wind_dir]
    draw_text(surf, f"Ветер: {arrow} {w.wind_dir}, {w.wind_speed:.0f} м/с",
              (20, y0 + 20), 16, COL_TEXT)
    draw_text(surf, f"Температура: {w.temp:.1f}°C",
              (20, y0 + 44), 16, COL_TEXT)
    draw_text(surf, f"Влажность: {w.humidity:.0f}%",
              (20, y0 + 68), 16, COL_TEXT)
    if w.rain_today > 0:
        draw_text(surf, f"Осадки: {w.rain_today:.1f} мм",
                  (20, y0 + 92), 16, COL_ACCENT2)
    else:
        draw_text(surf, "Осадков нет", (20, y0 + 92), 16, COL_TEXT_DIM)

    log_x = 340
    draw_text(surf, "ЖУРНАЛ СОБЫТИЙ", (log_x, y0), 12, COL_TEXT_DIM, bold=True)
    for i, msg in enumerate(list(game.events_log)[-6:]):
        draw_text(surf, msg, (log_x, y0 + 20 + i * 18), 14, COL_TEXT_DIM)

    hint_x = MAP_AREA_RIGHT - 320
    draw_text(surf, "УПРАВЛЕНИЕ", (hint_x, y0), 12, COL_TEXT_DIM, bold=True)
    draw_text(surf, f"Колёсико — зум (×{CAM_ZOOM:.2f})",
              (hint_x, y0 + 20), 13, COL_TEXT_DIM)
    draw_text(surf, "S — слоу-мо  |  P — вероятности",
              (hint_x, y0 + 40), 13, COL_TEXT_DIM)
    draw_text(surf, "ПКМ + перетаск. — панорама",
              (hint_x, y0 + 60), 13, COL_TEXT_DIM)
    draw_text(surf, "Space — следующий день",
              (hint_x, y0 + 80), 13, COL_TEXT_DIM)


def draw_tool_panel(surf, game):
    x0 = MAP_AREA_RIGHT
    pygame.draw.rect(surf, COL_PANEL, (x0, MAP_AREA_TOP, SCREEN_W - x0, MAP_AREA_H))
    pygame.draw.line(surf, COL_PANEL_HI, (x0, MAP_AREA_TOP), (x0, MAP_AREA_BOTTOM), 1)

    y = MAP_AREA_TOP + 20
    if game.phase == "build":
        draw_text(surf, "ФАЗА ПОДГОТОВКИ", (x0 + 16, y), 12, COL_ACCENT, bold=True)
        y += 18
        draw_text(surf, "Разместите средства защиты", (x0 + 16, y), 13, COL_TEXT_DIM)
        y += 28
    else:
        draw_text(surf, "ФАЗА СЕЗОНА", (x0 + 16, y), 12, COL_ACCENT, bold=True)
        y += 18
        if game.selected_unit is not None:
            draw_text(surf, "Выбран юнит — кликните цель", (x0 + 16, y), 13, COL_WARN)
        else:
            draw_text(surf, "Кликните на бригаду или базу", (x0 + 16, y), 13, COL_TEXT_DIM)
        y += 28

    buttons = []
    if game.phase == "build":
        for tool_key in [TOOL_WATCHTOWER, TOOL_SENSOR, TOOL_FIREBREAK, TOOL_BRIGADE, TOOL_PLANE]:
            t = TOOLS[tool_key]
            active = (game.selected_tool == tool_key)
            can_afford = (game.budget >= t["cost"])
            rect = pygame.Rect(x0 + 12, y, SCREEN_W - x0 - 24, 60)
            bg = COL_ACCENT if active else (COL_PANEL_HI if can_afford else darken(COL_PANEL_HI, 0.5))
            pygame.draw.rect(surf, bg, rect, border_radius=4)
            txt_col = (0, 0, 0) if active else (COL_TEXT if can_afford else COL_TEXT_DIM)
            draw_text(surf, t["name"], (rect.x + 10, rect.y + 8), 15, txt_col, bold=True)
            draw_text(surf, format_money(t["cost"]), (rect.right - 10, rect.y + 8), 13, txt_col, right=True)
            desc = t["desc"]
            if len(desc) > 40:
                split_at = desc.rfind(" ", 0, 40)
                if split_at > 0:
                    draw_text(surf, desc[:split_at], (rect.x + 10, rect.y + 28), 12, txt_col)
                    draw_text(surf, desc[split_at+1:], (rect.x + 10, rect.y + 42), 12, txt_col)
                else:
                    draw_text(surf, desc, (rect.x + 10, rect.y + 28), 12, txt_col)
            else:
                draw_text(surf, desc, (rect.x + 10, rect.y + 30), 12, txt_col)
            buttons.append(("tool", tool_key, rect))
            y += 68

    else:
        y += 8
        draw_text(surf, "РАЗМЕЩЕНО:", (x0 + 16, y), 12, COL_TEXT_DIM, bold=True)
        y += 22
        for i, p in enumerate(game.placements):
            t = TOOLS[p.kind]
            rect = pygame.Rect(x0 + 12, y, SCREEN_W - x0 - 24, 36)
            is_unit = p.kind in (TOOL_BRIGADE, TOOL_PLANE)
            sel = (game.selected_unit == i)
            bg = COL_ACCENT if sel else (COL_PANEL_HI if is_unit else COL_PANEL)
            pygame.draw.rect(surf, bg, rect, border_radius=3)
            txt_col = (0, 0, 0) if sel else COL_TEXT
            draw_text(surf, t["name"], (rect.x + 10, rect.y + 8), 13, txt_col, bold=is_unit)
            draw_text(surf, f"({p.x},{p.y})", (rect.right - 10, rect.y + 8), 12, txt_col, right=True)
            if is_unit:
                buttons.append(("unit", i, rect))
            y += 40

        # НОВОЕ: индикаторы режимов
        y += 8
        slow_mo_rect = pygame.Rect(x0 + 12, y, SCREEN_W - x0 - 24, 28)
        bg = COL_ACCENT if game.slow_mo else COL_PANEL_HI
        pygame.draw.rect(surf, bg, slow_mo_rect, border_radius=3)
        col = (0, 0, 0) if game.slow_mo else COL_TEXT
        draw_text(surf, f"[S] Слоу-мо: {'ВКЛ' if game.slow_mo else 'выкл'}",
                  (slow_mo_rect.x + 10, slow_mo_rect.y + 6), 13, col, bold=game.slow_mo)
        buttons.append(("toggle_slow", None, slow_mo_rect))
        y += 32

        prob_rect = pygame.Rect(x0 + 12, y, SCREEN_W - x0 - 24, 28)
        bg = COL_ACCENT if game.show_probabilities else COL_PANEL_HI
        pygame.draw.rect(surf, bg, prob_rect, border_radius=3)
        col = (0, 0, 0) if game.show_probabilities else COL_TEXT
        draw_text(surf, f"[P] Показ. вероятности: {'ВКЛ' if game.show_probabilities else 'выкл'}",
                  (prob_rect.x + 10, prob_rect.y + 6), 13, col, bold=game.show_probabilities)
        buttons.append(("toggle_prob", None, prob_rect))

    y_btn = MAP_AREA_BOTTOM - 110
    if game.phase == "build":
        rect = pygame.Rect(x0 + 12, y_btn, SCREEN_W - x0 - 24, 48)
        pygame.draw.rect(surf, COL_ACCENT, rect, border_radius=6)
        draw_text(surf, "НАЧАТЬ СЕЗОН ▶", rect.center, 18, (0, 0, 0), bold=True, center=True)
        buttons.append(("start", None, rect))
    elif game.phase == "play":
        rect = pygame.Rect(x0 + 12, y_btn, SCREEN_W - x0 - 24, 48)
        pygame.draw.rect(surf, COL_OK, rect, border_radius=6)
        draw_text(surf, "СЛЕДУЮЩИЙ ДЕНЬ ▶", rect.center, 18, (0, 0, 0), bold=True, center=True)
        buttons.append(("nextday", None, rect))

    y_btn += 56
    rect = pygame.Rect(x0 + 12, y_btn, SCREEN_W - x0 - 24, 34)
    pygame.draw.rect(surf, darken(COL_PANEL_HI), rect, border_radius=4)
    pygame.draw.rect(surf, COL_PANEL_HI, rect, 1, border_radius=4)
    draw_text(surf, "Главное меню", rect.center, 14, COL_TEXT_DIM, center=True)
    buttons.append(("menu", None, rect))

    return buttons


def draw_tile_tooltip(surf, game, gx, gy, mouse_pos, show_p=False):
    """Расширенный tooltip с показом вероятностей для соседей огня."""
    c = game.grid.cell(gx, gy)
    lines = [f"Клетка ({gx}, {gy})"]
    if c.state == CELL_EMPTY:
        lines.append("Луг — пожары редки")
    elif c.state == CELL_WATER:
        lines.append("Водоём — нужен для Бе-200")
    elif c.state == CELL_ROAD:
        lines.append("Дорога — источник возгораний")
    elif c.state == CELL_VILLAGE:
        lines.append("⚠ ПОСЁЛОК — защитить любой ценой")
    elif c.state == CELL_FIREBREAK:
        lines.append("Минер. полоса (p = 0.05)")
    elif c.state == CELL_FOREST:
        fd = FOREST_DATA[c.forest]
        lines.append(f"{fd['name']} — класс ПО: {fd['class']}")
        lines.append(f"p_base = {fd['p_base']:.2f}  |  горит {fd['burn_time']} дн.")
        if c.wet_turns > 0:
            lines.append(f"💧 Мокрая ещё {c.wet_turns} хода")
        # Если на этой клетке была вычислена p в последнем шаге, покажем
        if c.last_p > 0.001:
            lines.append(f"🎲 Последняя p поджига: {c.last_p:.3f}")
    elif c.state == CELL_BURNING:
        lines.append("🔥 ПОЖАР — остановить!")
        lines.append(f"Догорит через {c.burn_time} дн.")
        # ГЛАВНОЕ НОВОЕ: показываем расчёт вероятностей для соседей
        if game.day > 0:
            lines.append("─── РАСЧЁТ p ДЛЯ СОСЕДЕЙ ───")
            for nx, ny, dx, dy in game.grid.neighbors4(gx, gy):
                nc = game.grid.cell(nx, ny)
                if nc.state == CELL_FOREST:
                    res = compute_p_for_neighbor(game.grid, game.weather,
                                                  gx, gy, nx, ny)
                    if res:
                        pb, wf, hf, pt = res
                        dir_arrow = ""
                        if dx == 0 and dy == -1: dir_arrow = "↑"
                        elif dx == 1 and dy == 0: dir_arrow = "→"
                        elif dx == 0 and dy == 1: dir_arrow = "↓"
                        elif dx == -1 and dy == 0: dir_arrow = "←"
                        lines.append(f"{dir_arrow} p={pt:.2f} (база×ветер×влаж = {pb:.2f}×{wf:.2f}×{hf:.2f})")
    elif c.state == CELL_BURNT:
        lines.append("Гарь — регенерация > 20 лет")

    w = 0
    for ln in lines:
        img = font(13).render(ln, True, COL_TEXT)
        w = max(w, img.get_width())
    h = 10 + 18 * len(lines)
    w += 20
    mx, my = mouse_pos
    tx, ty = mx + 16, my + 16
    if tx + w > SCREEN_W:
        tx = mx - w - 16
    if ty + h > SCREEN_H:
        ty = my - h - 16
    bg = pygame.Surface((w, h), pygame.SRCALPHA)
    bg.fill((10, 15, 18, 230))
    surf.blit(bg, (tx, ty))
    pygame.draw.rect(surf, COL_PANEL_HI, (tx, ty, w, h), 1)
    for i, ln in enumerate(lines):
        col = COL_TEXT
        if "─" in ln:
            col = COL_ACCENT
        elif ln.startswith(("↑", "→", "↓", "←")):
            col = COL_ACCENT2
        draw_text(surf, ln, (tx + 10, ty + 8 + i * 18), 13, col)




# ═══════════════════════════════════════════════════════════════════
#   ЧАСТЬ 3: СЦЕНЫ
# ═══════════════════════════════════════════════════════════════════

_PARTICLES = [(random.randint(0, SCREEN_W), random.randint(0, SCREEN_H),
               random.uniform(0.2, 0.8), random.randint(1, 2)) for _ in range(90)]

def draw_atmosphere(surf, anim_t):
    surf.fill(COL_BG)
    for i in range(60):
        col = (20 + i // 5, 30 + i // 4, 40 + i // 4)
        pygame.draw.rect(surf, col, (0, i * 5, SCREEN_W, 5))
    for i, (x, y, spd, r) in enumerate(_PARTICLES):
        ny = (y + anim_t * spd * 10) % SCREEN_H
        col = (180, 195, 205)
        pygame.draw.circle(surf, col, (int(x), int(ny)), r)


def draw_menu(surf, anim_t):
    draw_atmosphere(surf, anim_t)

    cx = SCREEN_W // 2
    draw_text(surf, "ОГНЕННЫЙ ФРОНТ", (cx, 140), 72, COL_ACCENT, bold=True, center=True)
    draw_text(surf, "ЗАЩИТА ТАЙГИ", (cx, 210), 36, COL_TEXT, bold=True, center=True)
    draw_text(surf, "научная стратегия 8–9 класс · v2",
              (cx, 255), 16, COL_TEXT_DIM, center=True)

    phase = math.sin(anim_t * 2)
    pygame.draw.polygon(surf, COL_ACCENT, [
        (cx - 160, 140), (cx - 140, 100 + phase * 5), (cx - 120, 140)
    ])
    pygame.draw.polygon(surf, COL_ACCENT, [
        (cx + 120, 140), (cx + 140, 100 + phase * 5), (cx + 160, 140)
    ])

    buttons = []
    btn_w, btn_h = 360, 56
    y = 320
    labels = [
        ("start", "▶ НОВАЯ ИГРА", COL_ACCENT, (0, 0, 0)),
        ("lab", "🔬 ЛАБОРАТОРИЯ ПЕРКОЛЯЦИИ", COL_ACCENT2, (0, 0, 0)),
        ("brief", "📖 ЧТО ТАКОЕ ПОРОГ ПЕРКОЛЯЦИИ?", COL_PANEL_HI, COL_TEXT),
        ("quit", "ВЫХОД", COL_PANEL, COL_TEXT_DIM),
    ]
    for key, txt, bg, fg in labels:
        rect = pygame.Rect(cx - btn_w // 2, y, btn_w, btn_h)
        pygame.draw.rect(surf, bg, rect, border_radius=8)
        if bg not in (COL_ACCENT, COL_ACCENT2):
            pygame.draw.rect(surf, COL_PANEL_HI, rect, 1, border_radius=8)
        draw_text(surf, txt, rect.center, 18 if bg in (COL_ACCENT, COL_ACCENT2) else 16,
                  fg, bold=True, center=True)
        buttons.append((key, rect))
        y += btn_h + 14

    draw_text(surf, "Научная модель: клеточный автомат Бака–Ченя–Тангена (1990)",
              (cx, SCREEN_H - 40), 14, COL_TEXT_DIM, center=True)
    return buttons


def draw_briefing(surf, anim_t):
    draw_atmosphere(surf, anim_t)

    cx = SCREEN_W // 2
    y = 60
    draw_text(surf, "СОСНОВСКИЙ ЛЕСХОЗ — БРИФИНГ", (cx, y), 32, COL_ACCENT, bold=True, center=True)
    y += 50

    x_col = 120
    draw_text(surf, "ТВОЯ МИССИЯ", (x_col, y), 16, COL_ACCENT, bold=True)
    y2 = y + 28
    lines = [
        "• Защитить 2400 га тайги в течение 60 дней",
        "• Не допустить возгорания посёлка «Сосновка»",
        "• Стартовый бюджет: 50 млн ₽",
        "",
        "Ты располагаешь пятью инструментами защиты,",
        "каждый со своей стоимостью и ролью:",
        "",
        "— Вышка наблюдения: обнаружение в радиусе 8 км",
        "— Датчик дыма: точечное раннее предупреждение",
        "— Минерализованная полоса: физическая преграда",
        "— Наземная бригада: мобильное тушение",
        "— Бе-200: авиация, требует водоём",
        "",
        "В фазе подготовки расставляй защиту,",
        "в фазе сезона — реагируй на пожары каждый день.",
        "",
        "💡 НАЖМИ S — режим слоу-мо",
        "💡 НАЖМИ P — показать вероятности на карте",
    ]
    for ln in lines:
        col = COL_ACCENT if ln.startswith("💡") else COL_TEXT
        draw_text(surf, ln, (x_col, y2), 15, col)
        y2 += 22

    x_col2 = SCREEN_W // 2 + 60
    draw_text(surf, "НАУЧНАЯ СУТЬ", (x_col2, y), 16, COL_ACCENT, bold=True)
    y2 = y + 28
    lines2 = [
        "Под капотом — клеточный автомат Бака, Ченя",
        "и Тангеманна (1990). Каждая клетка — 1 гектар.",
        "",
        "Правила на каждый шаг (день):",
        "  ГОРИТ → догорает → пепелище",
        "  ЛЕС у ГОРЯЩЕГО  → загорается с вероят. p",
        "  ЛЕС + МОЛНИЯ → загорается редко (f)",
        "",
        "Вероятность p зависит от:",
        "  — типа леса (сосняк > ельник > березняк)",
        "  — ветра (по ветру ×1.8, против ×0.6)",
        "  — влажности (сухо ×1.6, влажно ×0.4)",
        "",
        "Индекс Нестерова (класс опасности):",
        "    ПВ = Σ T·(T − T_росы)",
        "",
        "Ключевая цифра: ρ_крит ≈ 0.59",
        "При плотности леса выше — пожар «проскакивает»",
        "всю карту. Это ПОРОГ ПЕРКОЛЯЦИИ.",
    ]
    for ln in lines2:
        col = COL_ACCENT2 if "ПВ =" in ln or "ρ_крит" in ln else COL_TEXT
        draw_text(surf, ln, (x_col2, y2), 15, col)
        y2 += 22

    buttons = []
    btn_w, btn_h = 240, 50
    rect1 = pygame.Rect(cx - btn_w - 20, SCREEN_H - 90, btn_w, btn_h)
    pygame.draw.rect(surf, COL_PANEL_HI, rect1, border_radius=6)
    draw_text(surf, "← Назад", rect1.center, 18, COL_TEXT, center=True)
    buttons.append(("back", rect1))

    rect2 = pygame.Rect(cx + 20, SCREEN_H - 90, btn_w, btn_h)
    pygame.draw.rect(surf, COL_ACCENT, rect2, border_radius=6)
    draw_text(surf, "Настроить игру →", rect2.center, 18, (0, 0, 0), bold=True, center=True)
    buttons.append(("go", rect2))

    return buttons


# ─── НОВОЕ: ЭКРАН НАСТРОЙКИ ИГРЫ ───
def draw_setup(surf, setup, anim_t):
    """Экран выбора параметров перед стартом игры."""
    draw_atmosphere(surf, anim_t)
    cx = SCREEN_W // 2
    y = 60

    draw_text(surf, "НАСТРОЙКИ МИССИИ", (cx, y), 32, COL_ACCENT, bold=True, center=True)
    y += 50
    draw_text(surf, "Эти параметры определяют сложность и характер сезона",
              (cx, y), 16, COL_TEXT_DIM, center=True)
    y += 50

    buttons = []
    panel_x = cx - 360
    panel_w = 720

    # Плотность леса (ползунок)
    draw_text(surf, "ПЛОТНОСТЬ ЛЕСА  ρ", (panel_x, y), 14, COL_ACCENT, bold=True)
    draw_text(surf, f"{setup.density:.2f}", (panel_x + panel_w, y), 14, COL_TEXT, bold=True, right=True)
    y += 24
    slider_rect = pygame.Rect(panel_x, y, panel_w, 20)
    pygame.draw.rect(surf, COL_PANEL_HI, slider_rect, border_radius=10)
    # Шкала с критическим порогом
    crit_x = panel_x + int(panel_w * 0.59)
    pygame.draw.line(surf, COL_DANGER, (crit_x, y - 4), (crit_x, y + 24), 2)
    draw_text(surf, "ρ_кр", (crit_x, y + 24), 11, COL_DANGER, center=False)
    # Ползунок
    knob_x = panel_x + int(panel_w * setup.density)
    knob_color = COL_ACCENT if setup.density < 0.59 else COL_DANGER
    pygame.draw.circle(surf, knob_color, (knob_x, y + 10), 11)
    pygame.draw.circle(surf, (0, 0, 0), (knob_x, y + 10), 11, 1)
    buttons.append(("density_slider", slider_rect, None))
    y += 42

    if setup.density < 0.5:
        hint = "🌲 Разреженный лес — пожары быстро гаснут"
        hint_col = COL_OK
    elif setup.density < 0.59:
        hint = "🌲 Близко к порогу — становится опасно"
        hint_col = COL_WARN
    else:
        hint = "🔥 Выше порога перколяции — огонь распространится повсюду!"
        hint_col = COL_DANGER
    draw_text(surf, hint, (panel_x, y), 13, hint_col, bold=True)
    y += 32

    # Тип карты (кнопки)
    draw_text(surf, "ТИП КАРТЫ", (panel_x, y), 14, COL_ACCENT, bold=True)
    y += 24
    map_btn_w = panel_w // 4 - 6
    for i, (key, mt) in enumerate(MAP_TYPES.items()):
        bx = panel_x + i * (map_btn_w + 8)
        rect = pygame.Rect(bx, y, map_btn_w, 60)
        active = (setup.map_type == key)
        bg = COL_ACCENT if active else COL_PANEL_HI
        pygame.draw.rect(surf, bg, rect, border_radius=6)
        col = (0, 0, 0) if active else COL_TEXT
        draw_text(surf, mt["name"], (rect.x + 8, rect.y + 8), 13, col, bold=True)
        # Описание в 2 строки
        desc = mt["desc"]
        draw_text(surf, desc[:32] if len(desc) > 32 else desc, (rect.x + 8, rect.y + 28), 11, col)
        if len(desc) > 32:
            draw_text(surf, desc[32:64], (rect.x + 8, rect.y + 42), 11, col)
        buttons.append(("map_type", rect, key))
    y += 74

    # Бюджет (ползунок)
    draw_text(surf, "СТАРТОВЫЙ БЮДЖЕТ", (panel_x, y), 14, COL_ACCENT, bold=True)
    draw_text(surf, format_money(setup.budget), (panel_x + panel_w, y), 14, COL_TEXT, bold=True, right=True)
    y += 24
    slider_rect = pygame.Rect(panel_x, y, panel_w, 20)
    pygame.draw.rect(surf, COL_PANEL_HI, slider_rect, border_radius=10)
    bmin, bmax = 10_000_000, 100_000_000
    bfrac = (setup.budget - bmin) / (bmax - bmin)
    knob_x = panel_x + int(panel_w * bfrac)
    pygame.draw.circle(surf, COL_ACCENT2, (knob_x, y + 10), 11)
    pygame.draw.circle(surf, (0, 0, 0), (knob_x, y + 10), 11, 1)
    buttons.append(("budget_slider", slider_rect, None))
    y += 42

    # Длительность сезона (ползунок)
    draw_text(surf, "ДЛИТЕЛЬНОСТЬ СЕЗОНА", (panel_x, y), 14, COL_ACCENT, bold=True)
    draw_text(surf, f"{setup.max_days} дней", (panel_x + panel_w, y), 14, COL_TEXT, bold=True, right=True)
    y += 24
    slider_rect = pygame.Rect(panel_x, y, panel_w, 20)
    pygame.draw.rect(surf, COL_PANEL_HI, slider_rect, border_radius=10)
    dfrac = (setup.max_days - 30) / (120 - 30)
    knob_x = panel_x + int(panel_w * dfrac)
    pygame.draw.circle(surf, COL_ACCENT2, (knob_x, y + 10), 11)
    pygame.draw.circle(surf, (0, 0, 0), (knob_x, y + 10), 11, 1)
    buttons.append(("days_slider", slider_rect, None))
    y += 42

    # Seed
    draw_text(surf, f"SEED ГЕНЕРАЦИИ: {setup.seed}", (panel_x, y), 13, COL_TEXT_DIM)
    rect = pygame.Rect(panel_x + 200, y - 4, 130, 26)
    pygame.draw.rect(surf, COL_PANEL_HI, rect, border_radius=4)
    draw_text(surf, "🎲 Случайный", rect.center, 12, COL_TEXT, center=True)
    buttons.append(("random_seed", rect, None))

    # Главные кнопки
    btn_w, btn_h = 240, 50
    rect1 = pygame.Rect(cx - btn_w - 20, SCREEN_H - 90, btn_w, btn_h)
    pygame.draw.rect(surf, COL_PANEL_HI, rect1, border_radius=6)
    draw_text(surf, "← Меню", rect1.center, 18, COL_TEXT, center=True)
    buttons.append(("back", rect1, None))

    rect2 = pygame.Rect(cx + 20, SCREEN_H - 90, btn_w, btn_h)
    pygame.draw.rect(surf, COL_ACCENT, rect2, border_radius=6)
    draw_text(surf, "Начать игру →", rect2.center, 18, (0, 0, 0), bold=True, center=True)
    buttons.append(("go", rect2, None))

    return buttons


# ─── НОВОЕ: ЛАБОРАТОРИЯ ПЕРКОЛЯЦИИ ───
def draw_lab(surf, lab, anim_t, hover_cell=None):
    """Главный экран лаборатории."""
    surf.fill(COL_BG)

    # Верхняя полоса
    pygame.draw.rect(surf, COL_PANEL, (0, 0, SCREEN_W, MAP_AREA_TOP))
    pygame.draw.line(surf, COL_PANEL_HI, (0, MAP_AREA_TOP), (SCREEN_W, MAP_AREA_TOP), 1)
    draw_text(surf, "🔬 ЛАБОРАТОРИЯ ПЕРКОЛЯЦИИ", (20, 8), 14, COL_ACCENT2, bold=True)
    draw_text(surf, "Чистый эксперимент: один поджог в центре, никаких юнитов", (20, 28), 16, COL_TEXT)

    # HUD статистики
    n_exp = len(lab.experiments)
    if n_exp > 0:
        last = lab.experiments[-1]
        col_x = 700
        draw_text(surf, "ОПЫТОВ", (col_x, 10), 11, COL_TEXT_DIM, bold=True)
        draw_text(surf, f"{n_exp}", (col_x, 26), 22, COL_TEXT, bold=True)
        col_x += 100
        draw_text(surf, "ПОСЛЕДНИЙ", (col_x, 10), 11, COL_TEXT_DIM, bold=True)
        col = COL_DANGER if last.burnt_frac > 0.5 else COL_OK
        draw_text(surf, f"{last.burnt_frac*100:.0f}%", (col_x, 26), 22, col, bold=True)
        col_x += 130
        draw_text(surf, "ρ ОПЫТА", (col_x, 10), 11, COL_TEXT_DIM, bold=True)
        draw_text(surf, f"{last.density:.2f}", (col_x, 26), 22, COL_ACCENT2, bold=True)

    if lab.batch_remaining > 0:
        col_x = 1100
        draw_text(surf, "ОЧЕРЕДЬ", (col_x, 10), 11, COL_ACCENT, bold=True)
        prog = lab.batch_total - lab.batch_remaining
        draw_text(surf, f"{prog} / {lab.batch_total}", (col_x, 26), 22, COL_ACCENT, bold=True)

    # Карта (или пустая область)
    if lab.grid is not None:
        draw_lab_map(surf, lab, anim_t, hover_cell)
    else:
        # Пустая зона с подсказкой
        cx = (MAP_AREA_LEFT + MAP_AREA_RIGHT) // 2
        cy = (MAP_AREA_TOP + MAP_AREA_BOTTOM) // 2
        draw_text(surf, "Задайте плотность ρ и нажмите «Один опыт»",
                  (cx, cy), 18, COL_TEXT_DIM, center=True)
        draw_text(surf, "или «Серия 100» для автопрогона",
                  (cx, cy + 30), 16, COL_TEXT_DIM, center=True)

    # Правая панель управления
    buttons = []
    panel_x = MAP_AREA_RIGHT
    pygame.draw.rect(surf, COL_PANEL, (panel_x, MAP_AREA_TOP, SCREEN_W - panel_x, MAP_AREA_H))
    pygame.draw.line(surf, COL_PANEL_HI, (panel_x, MAP_AREA_TOP), (panel_x, MAP_AREA_BOTTOM), 1)

    py = MAP_AREA_TOP + 16
    draw_text(surf, "ПАРАМЕТРЫ ОПЫТА", (panel_x + 16, py), 12, COL_ACCENT2, bold=True)
    py += 24

    # Плотность
    draw_text(surf, "Плотность ρ", (panel_x + 16, py), 13, COL_TEXT)
    draw_text(surf, f"{lab.density:.2f}",
              (panel_x + (SCREEN_W - panel_x) - 16, py), 14, COL_ACCENT, bold=True, right=True)
    py += 22
    sl_rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 16)
    pygame.draw.rect(surf, COL_PANEL_HI, sl_rect, border_radius=8)
    crit_x = sl_rect.x + int(sl_rect.w * 0.59)
    pygame.draw.line(surf, COL_DANGER, (crit_x, sl_rect.y), (crit_x, sl_rect.y + sl_rect.h), 2)
    knob_x = sl_rect.x + int(sl_rect.w * lab.density)
    pygame.draw.circle(surf, COL_ACCENT, (knob_x, sl_rect.y + 8), 9)
    buttons.append(("lab_density", sl_rect, None))
    py += 26

    # Скорость ветра
    draw_text(surf, "Ветер, м/с", (panel_x + 16, py), 13, COL_TEXT)
    draw_text(surf, f"{lab.wind_speed:.0f}",
              (panel_x + (SCREEN_W - panel_x) - 16, py), 14, COL_ACCENT, bold=True, right=True)
    py += 22
    sl_rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 16)
    pygame.draw.rect(surf, COL_PANEL_HI, sl_rect, border_radius=8)
    knob_x = sl_rect.x + int(sl_rect.w * lab.wind_speed / 25)
    pygame.draw.circle(surf, COL_ACCENT, (knob_x, sl_rect.y + 8), 9)
    buttons.append(("lab_wind", sl_rect, None))
    py += 26

    # Влажность
    draw_text(surf, "Влажность, %", (panel_x + 16, py), 13, COL_TEXT)
    draw_text(surf, f"{lab.humidity:.0f}",
              (panel_x + (SCREEN_W - panel_x) - 16, py), 14, COL_ACCENT, bold=True, right=True)
    py += 22
    sl_rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 16)
    pygame.draw.rect(surf, COL_PANEL_HI, sl_rect, border_radius=8)
    knob_x = sl_rect.x + int(sl_rect.w * lab.humidity / 100)
    pygame.draw.circle(surf, COL_ACCENT, (knob_x, sl_rect.y + 8), 9)
    buttons.append(("lab_humid", sl_rect, None))
    py += 32

    # Тип леса
    draw_text(surf, "Тип леса", (panel_x + 16, py), 13, COL_TEXT)
    py += 22
    btn_w = (SCREEN_W - panel_x - 32) // 2 - 4
    for i, (ftype, fd) in enumerate([(F_PINE, FOREST_DATA[F_PINE]),
                                      (F_BIRCH, FOREST_DATA[F_BIRCH])]):
        rect = pygame.Rect(panel_x + 16 + i * (btn_w + 8), py, btn_w, 30)
        active = (lab.forest_type == ftype)
        bg = COL_ACCENT if active else COL_PANEL_HI
        pygame.draw.rect(surf, bg, rect, border_radius=4)
        col = (0, 0, 0) if active else COL_TEXT
        draw_text(surf, fd["name"], rect.center, 13, col, bold=True, center=True)
        buttons.append(("lab_ftype", rect, ftype))
    py += 38

    # Кнопки управления
    py += 6
    draw_text(surf, "ЭКСПЕРИМЕНТ", (panel_x + 16, py), 12, COL_ACCENT2, bold=True)
    py += 22

    busy = lab.is_running or lab.batch_remaining > 0

    rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 36)
    if busy:
        pygame.draw.rect(surf, darken(COL_PANEL_HI, 0.6), rect, border_radius=4)
        draw_text(surf, "Один опыт (ожидание)", rect.center, 14, COL_TEXT_DIM, center=True)
    else:
        pygame.draw.rect(surf, COL_OK, rect, border_radius=4)
        draw_text(surf, "▶ Один опыт (с анимацией)", rect.center, 14, (0, 0, 0), bold=True, center=True)
        buttons.append(("lab_run_one", rect, None))
    py += 42

    rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 36)
    if busy:
        pygame.draw.rect(surf, darken(COL_PANEL_HI, 0.6), rect, border_radius=4)
        draw_text(surf, "Серия 100 (ожидание)", rect.center, 14, COL_TEXT_DIM, center=True)
    else:
        pygame.draw.rect(surf, COL_ACCENT, rect, border_radius=4)
        draw_text(surf, "⚡ 100 опытов (быстро)", rect.center, 14, (0, 0, 0), bold=True, center=True)
        buttons.append(("lab_batch_fast", rect, None))
    py += 42

    rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 36)
    if busy:
        pygame.draw.rect(surf, darken(COL_PANEL_HI, 0.6), rect, border_radius=4)
        draw_text(surf, "Серия 100 (ожидание)", rect.center, 14, COL_TEXT_DIM, center=True)
    else:
        pygame.draw.rect(surf, COL_ACCENT2, rect, border_radius=4)
        draw_text(surf, "📺 100 опытов (с показом)", rect.center, 14, (0, 0, 0), bold=True, center=True)
        buttons.append(("lab_batch_animate", rect, None))
    py += 42

    if busy:
        rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 30)
        pygame.draw.rect(surf, COL_DANGER, rect, border_radius=4)
        draw_text(surf, "■ Стоп", rect.center, 14, COL_TEXT, bold=True, center=True)
        buttons.append(("lab_stop", rect, None))
        py += 36

    py = MAP_AREA_BOTTOM - 130
    rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 30)
    pygame.draw.rect(surf, COL_PANEL_HI, rect, border_radius=4)
    draw_text(surf, "🗑 Очистить опыты", rect.center, 13, COL_TEXT, center=True)
    buttons.append(("lab_clear", rect, None))
    py += 36

    rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 30)
    pygame.draw.rect(surf, COL_ACCENT2, rect, border_radius=4)
    draw_text(surf, "📊 Экспорт в Excel", rect.center, 13, (0, 0, 0), bold=True, center=True)
    buttons.append(("lab_export", rect, None))
    py += 36

    rect = pygame.Rect(panel_x + 16, py, SCREEN_W - panel_x - 32, 30)
    pygame.draw.rect(surf, darken(COL_PANEL_HI), rect, border_radius=4)
    draw_text(surf, "← Главное меню", rect.center, 13, COL_TEXT_DIM, center=True)
    buttons.append(("lab_menu", rect, None))

    # ─── ГРАФИК ВНИЗУ ───
    draw_lab_chart(surf, lab)

    return buttons


def draw_lab_chart(surf, lab):
    """Точечный график «плотность vs % сгоревшего» в нижней панели."""
    chart_x = 0
    chart_y = MAP_AREA_BOTTOM
    chart_w = MAP_AREA_RIGHT
    chart_h = SCREEN_H - MAP_AREA_BOTTOM

    pygame.draw.rect(surf, COL_PANEL, (chart_x, chart_y, chart_w, chart_h))
    pygame.draw.line(surf, COL_PANEL_HI, (0, chart_y), (chart_w, chart_y), 1)

    pad_l = 80
    pad_r = 30
    pad_t = 28
    pad_b = 30
    plot_x0 = chart_x + pad_l
    plot_y0 = chart_y + pad_t
    plot_w = chart_w - pad_l - pad_r
    plot_h = chart_h - pad_t - pad_b

    # Заголовок
    draw_text(surf, "ГРАФИК ЭКСПЕРИМЕНТОВ: плотность ρ → доля сгоревшего, %",
              (chart_x + 14, chart_y + 6), 13, COL_ACCENT2, bold=True)

    # Оси
    pygame.draw.line(surf, COL_TEXT_DIM, (plot_x0, plot_y0),
                     (plot_x0, plot_y0 + plot_h), 1)
    pygame.draw.line(surf, COL_TEXT_DIM, (plot_x0, plot_y0 + plot_h),
                     (plot_x0 + plot_w, plot_y0 + plot_h), 1)

    # Подписи осей Y
    for frac, label in [(0, "0"), (0.25, "25"), (0.5, "50"), (0.75, "75"), (1.0, "100%")]:
        py = plot_y0 + plot_h - int(plot_h * frac)
        draw_text(surf, label, (plot_x0 - 6, py - 6), 11, COL_TEXT_DIM, right=True)
        pygame.draw.line(surf, darken(COL_TEXT_DIM, 0.4),
                         (plot_x0, py), (plot_x0 + plot_w, py), 1)

    # Подписи X
    for frac, label in [(0, "0.0"), (0.25, "0.25"), (0.5, "0.5"), (0.75, "0.75"), (1.0, "1.0")]:
        px = plot_x0 + int(plot_w * frac)
        draw_text(surf, label, (px, plot_y0 + plot_h + 4), 11, COL_TEXT_DIM, center=False)

    # Линия порога перколяции
    crit_x = plot_x0 + int(plot_w * 0.59)
    for i in range(0, plot_h, 6):
        pygame.draw.line(surf, COL_DANGER,
                         (crit_x, plot_y0 + i), (crit_x, plot_y0 + i + 3), 1)
    draw_text(surf, "ρ_кр = 0.59", (crit_x + 4, plot_y0), 11, COL_DANGER, bold=True)

    # Точки экспериментов
    for exp in lab.experiments:
        px = plot_x0 + int(plot_w * exp.density)
        py = plot_y0 + plot_h - int(plot_h * exp.burnt_frac)
        col = COL_DANGER if exp.burnt_frac > 0.5 else COL_OK
        pygame.draw.circle(surf, col, (px, py), 3)

    # Текущая позиция (где будет следующий опыт)
    if not lab.is_running and lab.batch_remaining == 0:
        px = plot_x0 + int(plot_w * lab.density)
        for i in range(0, plot_h, 4):
            pygame.draw.line(surf, darken(COL_ACCENT, 0.7),
                             (px, plot_y0 + i), (px, plot_y0 + i + 2), 1)


# ─── НОВОЕ: ИТОГОВЫЙ РАЗБОР (немного обновлён) ───
def draw_analysis(surf, game, anim_t):
    draw_atmosphere(surf, anim_t)
    cx = SCREEN_W // 2
    y = 40

    score, saved_frac, shannon, tier = game.compute_score()

    if game.end_reason == "village":
        draw_text(surf, "МИССИЯ ПРОВАЛЕНА", (cx, y), 44, COL_DANGER, bold=True, center=True)
        draw_text(surf, "Посёлок Сосновка охвачен огнём", (cx, y + 50), 20, COL_TEXT_DIM, center=True)
    else:
        color = {"GOLD": COL_WARN, "SILVER": COL_TEXT, "BRONZE": (200, 140, 80), "—": COL_TEXT_DIM}[tier]
        draw_text(surf, "СЕЗОН ЗАВЕРШЁН", (cx, y), 44, COL_OK, bold=True, center=True)
        if tier != "—":
            draw_text(surf, f"Результат: {tier}", (cx, y + 50), 24, color, bold=True, center=True)
        else:
            draw_text(surf, "Результат: без награды", (cx, y + 50), 20, COL_TEXT_DIM, center=True)

    y0 = 150
    x_col = 100
    g = game.grid
    total_initial = sum(g.initial_forest_ha.values())
    draw_text(surf, "ИТОГОВАЯ СТАТИСТИКА", (x_col, y0), 16, COL_ACCENT, bold=True)
    rows = [
        ("Дней сыграно", f"{game.day}"),
        ("Исходный лесной покров", f"{total_initial} га"),
        ("Сгорело всего", f"{g.total_ha_burnt} га ({100*g.total_ha_burnt/max(total_initial,1):.1f}%)"),
        ("Лес сохранён", f"{g.remaining_forest_ha()} га ({saved_frac*100:.1f}%)"),
        ("Потрачено средств", f"{(game.initial_budget-game.budget)/1_000_000:.1f} млн ₽"),
        ("Индекс Шеннона", f"{shannon:.3f}"),
        ("Выброшено CO₂", f"{int(sum(g.burnt_by_type.get(f, 0) * FOREST_DATA[f]['biomass'] * 0.47 for f in [F_PINE, F_SPRUCE, F_BIRCH, F_PEAT]))} т"),
        ("Активных очагов к концу", f"{game.active_fires}"),
        ("Итоговый балл", f"{score:.0f}"),
    ]
    y1 = y0 + 30
    for label, val in rows:
        draw_text(surf, label, (x_col, y1), 14, COL_TEXT_DIM)
        draw_text(surf, val, (x_col + 400, y1), 16, COL_TEXT, bold=True)
        y1 += 28

    x_chart = SCREEN_W // 2 + 60
    draw_text(surf, "ПОТЕРИ ПО ТИПАМ ЛЕСА", (x_chart, y0), 16, COL_ACCENT, bold=True)
    y1 = y0 + 40
    bar_w = 400
    for f in [F_PINE, F_SPRUCE, F_BIRCH, F_PEAT]:
        fd = FOREST_DATA[f]
        lost = g.burnt_by_type.get(f, 0)
        init = g.initial_forest_ha.get(f, 1)
        pct = lost / init if init > 0 else 0
        draw_text(surf, f"{fd['name']} (класс {fd['class']})", (x_chart, y1), 14, COL_TEXT)
        draw_text(surf, f"{lost} / {init} га", (x_chart + bar_w, y1), 13, COL_TEXT_DIM, right=True)
        pygame.draw.rect(surf, darken(fd["color"], 0.5), (x_chart, y1 + 22, bar_w, 16))
        fill_w = int(bar_w * pct)
        pygame.draw.rect(surf, fd["color"], (x_chart, y1 + 22, fill_w, 16))
        pygame.draw.rect(surf, COL_PANEL_HI, (x_chart, y1 + 22, bar_w, 16), 1)
        draw_text(surf, f"{pct*100:.0f}% потеряно", (x_chart + bar_w // 2, y1 + 22 + 2),
                  12, (255, 255, 255), bold=True, center=False)
        y1 += 58

    y1 += 16
    draw_text(surf, "НАУЧНЫЙ ВЫВОД", (x_chart, y1), 16, COL_ACCENT, bold=True)
    y1 += 26
    density = game.setup.density
    draw_text(surf, f"Стартовая плотность леса: {density:.2f}", (x_chart, y1), 14, COL_TEXT)
    y1 += 22
    if density > 0.59:
        draw_text(surf, f"⚠ Выше порога перколяции (0.59)",
                  (x_chart, y1), 14, COL_DANGER, bold=True)
        y1 += 22
        draw_text(surf, "Огонь имел шанс распространиться на всю карту.",
                  (x_chart, y1), 13, COL_TEXT_DIM)
    else:
        draw_text(surf, f"✓ Ниже порога перколяции (0.59)",
                  (x_chart, y1), 14, COL_OK, bold=True)
        y1 += 22
        draw_text(surf, "При такой плотности огонь склонен гаснуть сам.",
                  (x_chart, y1), 13, COL_TEXT_DIM)

    buttons = []
    btn_w, btn_h = 280, 54
    rect1 = pygame.Rect(cx - btn_w - 160, SCREEN_H - 80, btn_w, btn_h)
    pygame.draw.rect(surf, COL_ACCENT2, rect1, border_radius=6)
    draw_text(surf, "📊 ЭКСПОРТ В EXCEL", rect1.center, 18, (0, 0, 0), bold=True, center=True)
    buttons.append(("export", rect1))

    rect2 = pygame.Rect(cx - btn_w // 2 + 160, SCREEN_H - 80, btn_w, btn_h)
    pygame.draw.rect(surf, COL_PANEL_HI, rect2, border_radius=6)
    draw_text(surf, "ГЛАВНОЕ МЕНЮ", rect2.center, 18, COL_TEXT, bold=True, center=True)
    buttons.append(("menu", rect2))

    return buttons



# ═══════════════════════════════════════════════════════════════════
#   ЧАСТЬ 4: ЭКСПОРТ В EXCEL
# ═══════════════════════════════════════════════════════════════════

FONT_NAME = "Arial"

def style_header(cell):
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2F6E4A")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_subheader(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True)
    cell.fill = PatternFill("solid", fgColor="E8EFE3")

def style_title(cell):
    cell.font = Font(name=FONT_NAME, size=14, bold=True, color="1F4530")
    cell.alignment = Alignment(horizontal="left", vertical="center")

def style_body(cell, bold=False):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold)

def apply_borders(ws, cell_range):
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[cell_range]:
        for c in row:
            c.border = border


def sheet_summary(wb, game):
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "ОГНЕННЫЙ ФРОНТ — СВОДКА ПО МИССИИ"
    ws.merge_cells("A1:D1")
    style_title(ws["A1"])

    ws["A3"] = "Дата игры"
    ws["B3"] = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws["A4"] = "Seed карты"
    ws["B4"] = game.seed
    ws["A5"] = "Плотность леса ρ"
    ws["B5"] = round(game.setup.density, 3)
    ws["A6"] = "Тип карты"
    ws["B6"] = MAP_TYPES[game.setup.map_type]["name"]
    ws["A7"] = "Причина завершения"
    reason_map = {"days": "Сезон завершён", "village": "Посёлок охвачен", "quit": "Выход"}
    ws["B7"] = reason_map.get(game.end_reason, "—")

    ws["A9"] = "БЮДЖЕТ"
    ws.merge_cells("A9:B9")
    style_subheader(ws["A9"])
    ws["A10"] = "Стартовый"
    ws["B10"] = game.initial_budget
    ws["A11"] = "Потрачено"
    ws["B11"] = "=B10-B12"
    ws["A12"] = "Остаток"
    ws["B12"] = game.budget
    for c in ["B10", "B11", "B12"]:
        ws[c].number_format = '#,##0 "₽"'

    ws["A14"] = "ЛЕСНОЙ ПОКРОВ"
    ws.merge_cells("A14:B14")
    style_subheader(ws["A14"])
    g = game.grid
    total_initial = sum(g.initial_forest_ha.values())
    ws["A15"] = "Исходная площадь леса, га"
    ws["B15"] = total_initial
    ws["A16"] = "Сгорело, га"
    ws["B16"] = g.total_ha_burnt
    ws["A17"] = "Сохранено, га"
    ws["B17"] = "=B15-B16"
    ws["A18"] = "Доля сгоревшего, %"
    ws["B18"] = "=B16/B15*100"
    ws["B18"].number_format = '0.0"%"'

    ws["A20"] = "ОЦЕНКА"
    ws.merge_cells("A20:B20")
    style_subheader(ws["A20"])
    score, saved_frac, shannon, tier = game.compute_score()
    ws["A21"] = "Индекс Шеннона"
    ws["B21"] = round(shannon, 3)
    ws["A22"] = "Выброс CO₂, т"
    co2 = sum(g.burnt_by_type.get(f, 0) * FOREST_DATA[f]["biomass"] * 0.47
              for f in [F_PINE, F_SPRUCE, F_BIRCH, F_PEAT])
    ws["B22"] = int(co2)
    ws["A23"] = "Итоговый балл"
    ws["B23"] = round(score, 1)
    ws["A24"] = "Уровень"
    ws["B24"] = tier

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22


def sheet_timeline(wb, game):
    ws = wb.create_sheet("Хронология")
    headers = ["День", "Темп. °C", "Влажн. %", "Ветер", "Скор. м/с", "Осадки мм",
               "Индекс Нестерова", "Очагов", "Сгорело нараст., га",
               "Лес остаток, га", "Бюджет ₽"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        style_header(c)

    for r, day_data in enumerate(game.history, start=2):
        ws.cell(row=r, column=1, value=day_data["day"])
        ws.cell(row=r, column=2, value=day_data["temp"])
        ws.cell(row=r, column=3, value=day_data["humidity"])
        ws.cell(row=r, column=4, value=day_data["wind_dir"])
        ws.cell(row=r, column=5, value=day_data["wind_speed"])
        ws.cell(row=r, column=6, value=day_data["rain"])
        ws.cell(row=r, column=7, value=day_data["nesterov"])
        ws.cell(row=r, column=8, value=day_data["active_fires"])
        ws.cell(row=r, column=9, value=day_data["burnt_total"])
        ws.cell(row=r, column=10, value=day_data["forest_remaining"])
        ws.cell(row=r, column=11, value=day_data["budget"])

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 14

    n_rows = len(game.history) + 1
    if n_rows > 1:
        apply_borders(ws, f"A1:K{n_rows}")
        chart = LineChart()
        chart.title = "Площадь пожара по дням (га)"
        chart.style = 2
        chart.height = 10
        chart.width = 22
        data = Reference(ws, min_col=9, min_row=1, max_row=n_rows, max_col=9)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "M2")


def sheet_forest_types(wb, game):
    ws = wb.create_sheet("Типы леса")
    g = game.grid
    headers = ["Тип леса", "Класс ПО", "p_base", "Время горения",
               "Биомасса т/га", "Исходно, га", "Сгорело, га", "Потеряно %", "CO₂, т"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        style_header(c)
    row = 2
    for f in [F_PINE, F_SPRUCE, F_BIRCH, F_PEAT]:
        fd = FOREST_DATA[f]
        initial = g.initial_forest_ha.get(f, 0)
        burnt = g.burnt_by_type.get(f, 0)
        ws.cell(row=row, column=1, value=fd["name"])
        ws.cell(row=row, column=2, value=fd["class"])
        ws.cell(row=row, column=3, value=fd["p_base"])
        ws.cell(row=row, column=4, value=fd["burn_time"])
        ws.cell(row=row, column=5, value=fd["biomass"])
        ws.cell(row=row, column=6, value=initial)
        ws.cell(row=row, column=7, value=burnt)
        ws.cell(row=row, column=8, value=f"=IF(F{row}>0,G{row}/F{row}*100,0)")
        ws.cell(row=row, column=9, value=f"=G{row}*E{row}*0.47")
        row += 1
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 16


def sheet_lab_experiments(wb, lab):
    """Лист с данными лабораторных экспериментов — главный новый лист!"""
    ws = wb.create_sheet("Лаборатория", 0)  # делаем активным
    ws["A1"] = "ЛАБОРАТОРИЯ ПЕРКОЛЯЦИИ — ДАННЫЕ ОПЫТОВ"
    ws.merge_cells("A1:H1")
    style_title(ws["A1"])

    ws["A3"] = "Цель эксперимента: найти порог перколяции для ваших условий"
    ws["A4"] = "Гипотеза: при ρ > 0.59 на квадратной решётке огонь распространится на всю карту."
    ws["A5"] = "Меняй плотность ρ и наблюдай долю сгоревшего. Постройте по точкам S-кривую."

    headers = ["№", "ρ (плотность)", "Сгорело, %", "Сгорело, га",
               "Исходно, га", "Ветер, м/с", "Влажность, %", "Шагов", "Seed", "Время"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=7, column=i, value=h)
        style_header(c)

    for i, exp in enumerate(lab.experiments, start=8):
        ws.cell(row=i, column=1, value=i - 7)
        ws.cell(row=i, column=2, value=round(exp.density, 3))
        ws.cell(row=i, column=3, value=round(exp.burnt_frac * 100, 1))
        ws.cell(row=i, column=4, value=exp.total_burnt)
        ws.cell(row=i, column=5, value=exp.total_initial)
        ws.cell(row=i, column=6, value=round(exp.wind_speed, 1))
        ws.cell(row=i, column=7, value=round(exp.humidity, 1))
        ws.cell(row=i, column=8, value=exp.steps)
        ws.cell(row=i, column=9, value=exp.seed)
        ws.cell(row=i, column=10, value=exp.timestamp)

    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 14

    n_rows = len(lab.experiments)
    if n_rows > 0:
        apply_borders(ws, f"A7:J{7 + n_rows}")

        # Точечный график
        chart = ScatterChart()
        chart.title = "Доля сгоревшего vs плотность леса (поиск порога перколяции)"
        chart.style = 13
        chart.x_axis.title = "Плотность ρ"
        chart.y_axis.title = "Сгорело, %"
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 1
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.height = 14
        chart.width = 22

        x_ref = Reference(ws, min_col=2, min_row=8, max_row=7 + n_rows)
        y_ref = Reference(ws, min_col=3, min_row=8, max_row=7 + n_rows)
        series = Series(y_ref, x_ref, title="Эксперименты")
        chart.series.append(series)
        ws.add_chart(chart, "L7")

    # Теоретическая S-кривая для сравнения
    ws["L4"] = "ТЕОРЕТИЧЕСКАЯ КРИВАЯ"
    style_subheader(ws["L4"])
    ws["L5"] = "ρ"
    ws["M5"] = "Доля, %"
    style_header(ws["L5"])
    style_header(ws["M5"])
    densities = [round(0.05 * i, 2) for i in range(1, 20)]
    for i, d in enumerate(densities):
        ws.cell(row=6 + i, column=12, value=d)
        expected = 100 / (1 + math.exp(-60 * (d - 0.59)))
        ws.cell(row=6 + i, column=13, value=round(expected, 1))


def sheet_formulas(wb, game_or_lab):
    ws = wb.create_sheet("Формулы")
    ws["A1"] = "НАУЧНЫЕ ФОРМУЛЫ И ИСТОЧНИКИ"
    ws.merge_cells("A1:C1")
    style_title(ws["A1"])
    rows = [
        ("", "", ""),
        ("КЛЕТОЧНЫЙ АВТОМАТ (Бак, Чен, Тангеманн, 1990)", "", ""),
        ("Состояния клетки", "EMPTY, FOREST, BURNING, BURNT", ""),
        ("Переход ЛЕС → ГОРИТ", "P(ignite) = p_base × w × h × f", "для каждого соседа"),
        ("  p_base", "база по типу леса", "Мелехов, 1965"),
        ("  w (ветер)", "1 + 0.08 · v · cos(θ)", "θ — угол ветер↔направление"),
        ("  h (влажность)", "1.6 при <30%, 0.4 при >70%", ""),
        ("Длительность горения", "burn_time клеток, затем BURNT", ""),
        ("", "", ""),
        ("ИНДЕКС ГОРИМОСТИ ПО НЕСТЕРОВУ", "", ""),
        ("Формула", "ПВ = Σ T·(T−T_росы)", "за сухие дни"),
        ("Класс I—V", "ПВ < 300 — низкая, ≥ 10000 — чрезв.", ""),
        ("", "", ""),
        ("ПОРОГ ПЕРКОЛЯЦИИ", "", ""),
        ("Сайт-перколяция, квадратная решётка", "ρ_кр ≈ 0.5927", "Ньюман, 2000"),
        ("Физический смысл", "при ρ > ρ_кр существует бесконечный кластер", ""),
        ("", "", ""),
        ("ИНДЕКС ШЕННОНА", "H = −Σ p_i·ln(p_i)", "0 = монокультура"),
        ("ВЫБРОС CO₂", "M = S·B·0.47", "S — площадь, B — биомасса"),
        ("", "", ""),
        ("ИСТОЧНИКИ", "", ""),
        ("Bak, Chen, Tang. Phys.Lett.A. 1990.", "", ""),
        ("Мелехов И.С. Лесная пирология. 1965.", "", ""),
        ("Нестеров В.Г. Горимость леса. 1949.", "", ""),
        ("Newman M.E.J. Ziff R.M. Phys.Rev.E. 2001.", "", ""),
    ]
    for i, (a, b, c) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=3, value=c)
        if a and not b and not c:
            style_subheader(ws.cell(row=i, column=1))
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 35


def export_game(game, path=None):
    if not HAS_OPENPYXL:
        return None, "openpyxl не установлен"
    if path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            _, _, _, tier = game.compute_score()
            tier_suffix = f"_{tier}" if tier != "—" else ""
        except:
            tier_suffix = ""
        path = f"ОгненныйФронт_{ts}{tier_suffix}.xlsx"
    wb = Workbook()
    sheet_summary(wb, game)
    sheet_timeline(wb, game)
    sheet_forest_types(wb, game)
    sheet_formulas(wb, game)
    try:
        wb.save(path)
        return path, None
    except Exception as e:
        return None, str(e)


def export_lab(lab, path=None):
    """Экспорт лабораторных опытов в Excel."""
    if not HAS_OPENPYXL:
        return None, "openpyxl не установлен"
    if path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = f"Лаборатория_{ts}.xlsx"
    wb = Workbook()
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    sheet_lab_experiments(wb, lab)
    sheet_formulas(wb, lab)
    try:
        wb.save(path)
        return path, None
    except Exception as e:
        return None, str(e)


# ═══════════════════════════════════════════════════════════════════
#   ЧАСТЬ 5: ГЛАВНЫЙ ЦИКЛ
# ═══════════════════════════════════════════════════════════════════

def bresenham_line(x0, y0, x1, y1):
    points = []
    dx = abs(x1 - x0)
    dy = abs(y1 - y0)
    sx = 1 if x0 < x1 else -1
    sy = 1 if y0 < y1 else -1
    err = dx - dy
    x, y = x0, y0
    while True:
        points.append((x, y))
        if x == x1 and y == y1:
            break
        e2 = 2 * err
        if e2 > -dy:
            err -= dy
            x += sx
        if e2 < dx:
            err += dx
            y += sy
    return points


def main():
    global CAM_PAN_X, CAM_PAN_Y

    screen = pygame.display.set_mode((SCREEN_W, SCREEN_H))
    pygame.display.set_caption("Огненный фронт: Защита тайги — v2")
    clock = pygame.time.Clock()

    setup = GameSetup()
    game = GameState(setup)
    lab = PercolationLab()
    anim_t = 0.0
    running = True

    last_export_msg = None
    last_export_time = 0

    rmb_down_pos = None
    rmb_drag_active = False
    RMB_DRAG_THRESHOLD = 5

    # Для drag-обработки ползунков
    dragging_slider = None  # (kind, rect) — какой ползунок тянем

    while running:
        dt = clock.tick(60) / 1000.0
        anim_t += dt
        mouse_pos = pygame.mouse.get_pos()
        mouse_pressed = pygame.mouse.get_pressed()

        # Обновление лаборатории каждый кадр
        if game.phase == "lab":
            lab.update_batch()

        # В слоу-мо в основной игре — авто-проигрывание дней раз в N мс
        # (только если нет активных пожаров: иначе даём игроку реагировать)

        hover_cell = None
        if game.phase in ("build", "play"):
            if MAP_AREA_LEFT <= mouse_pos[0] < MAP_AREA_RIGHT and MAP_AREA_TOP <= mouse_pos[1] < MAP_AREA_BOTTOM:
                hover_cell = screen_to_grid(*mouse_pos)
        elif game.phase == "lab":
            if MAP_AREA_LEFT <= mouse_pos[0] < MAP_AREA_RIGHT and MAP_AREA_TOP <= mouse_pos[1] < MAP_AREA_BOTTOM:
                hover_cell = screen_to_grid(*mouse_pos)

        # ─── ОБРАБОТКА drag-ползунков ───
        if dragging_slider is not None and mouse_pressed[0]:
            kind, rect = dragging_slider
            frac = max(0, min(1, (mouse_pos[0] - rect.x) / rect.w))
            if kind == "density_slider":
                setup.density = round(0.1 + frac * 0.85, 2)
            elif kind == "budget_slider":
                bmin, bmax = 10_000_000, 100_000_000
                setup.budget = int(bmin + frac * (bmax - bmin))
                setup.budget = (setup.budget // 1_000_000) * 1_000_000
            elif kind == "days_slider":
                setup.max_days = 30 + int(frac * 90)
            elif kind == "lab_density":
                lab.density = round(0.1 + frac * 0.85, 2)
            elif kind == "lab_wind":
                lab.wind_speed = round(frac * 25)
            elif kind == "lab_humid":
                lab.humidity = round(frac * 100)
        elif not mouse_pressed[0]:
            dragging_slider = None

        events = pygame.event.get()
        for ev in events:
            if ev.type == pygame.QUIT:
                running = False

            elif ev.type == pygame.KEYDOWN:
                if ev.key == pygame.K_ESCAPE:
                    if game.phase in ("build", "play"):
                        game.selected_tool = None
                        game.selected_unit = None
                        game.firebreak_anchor = None
                        game.plane_step = None
                    elif game.phase in ("brief", "analysis", "setup", "lab"):
                        game.phase = "menu"
                elif ev.key == pygame.K_SPACE and game.phase == "play":
                    game.advance_day()
                elif ev.key == pygame.K_RETURN and game.phase == "build":
                    game.start_season()
                elif ev.key in (pygame.K_HOME, pygame.K_0):
                    if game.phase in ("build", "play", "lab"):
                        camera_reset()
                elif ev.key in (pygame.K_PLUS, pygame.K_EQUALS, pygame.K_KP_PLUS):
                    if game.phase in ("build", "play", "lab"):
                        apply_zoom(CAM_ZOOM * 1.15, *mouse_pos)
                elif ev.key in (pygame.K_MINUS, pygame.K_KP_MINUS):
                    if game.phase in ("build", "play", "lab"):
                        apply_zoom(CAM_ZOOM / 1.15, *mouse_pos)
                # НОВОЕ: тогглы режимов
                elif ev.key == pygame.K_s and game.phase == "play":
                    game.slow_mo = not game.slow_mo
                    game._log(f"Слоу-мо: {'ВКЛ' if game.slow_mo else 'выкл'}")
                elif ev.key == pygame.K_p and game.phase == "play":
                    game.show_probabilities = not game.show_probabilities
                    game._log(f"Показ вероятностей: {'ВКЛ' if game.show_probabilities else 'выкл'}")
                # Лаборатория — пробел = один опыт
                elif ev.key == pygame.K_SPACE and game.phase == "lab":
                    if not lab.is_running and lab.batch_remaining == 0:
                        lab.setup_experiment(animate=True)

            elif ev.type == pygame.MOUSEWHEEL:
                if game.phase in ("build", "play", "lab"):
                    if (MAP_AREA_LEFT <= mouse_pos[0] < MAP_AREA_RIGHT and
                        MAP_AREA_TOP <= mouse_pos[1] < MAP_AREA_BOTTOM):
                        factor = 1.15 if ev.y > 0 else (1 / 1.15)
                        apply_zoom(CAM_ZOOM * factor, mouse_pos[0], mouse_pos[1])

            elif ev.type == pygame.MOUSEBUTTONDOWN and ev.button == 3:
                if game.phase in ("build", "play", "lab"):
                    rmb_down_pos = ev.pos
                    rmb_drag_active = False

            elif ev.type == pygame.MOUSEMOTION:
                if rmb_down_pos is not None and (ev.buttons[2] == 1):
                    dx = ev.pos[0] - rmb_down_pos[0]
                    dy = ev.pos[1] - rmb_down_pos[1]
                    if not rmb_drag_active:
                        if abs(dx) > RMB_DRAG_THRESHOLD or abs(dy) > RMB_DRAG_THRESHOLD:
                            rmb_drag_active = True
                    if rmb_drag_active:
                        CAM_PAN_X += ev.rel[0]
                        CAM_PAN_Y += ev.rel[1]

            elif ev.type == pygame.MOUSEBUTTONUP and ev.button == 3:
                if game.phase in ("build", "play"):
                    if not rmb_drag_active:
                        game.selected_tool = None
                        game.selected_unit = None
                        game.firebreak_anchor = None
                        game.plane_step = None
                rmb_down_pos = None
                rmb_drag_active = False

        screen.fill(COL_BG)

        # ─── ГЛАВНОЕ МЕНЮ ───
        if game.phase == "menu":
            btns = draw_menu(screen, anim_t)
            for ev in events:
                if ev.type == pygame.MOUSEBUTTONDOWN and ev.button == 1:
                    for key, rect in btns:
                        if rect.collidepoint(ev.pos):
                            if key == "start":
                                setup = GameSetup()
                                setup.seed = pygame.time.get_ticks() & 0xFFFF
                                game = GameState(setup)
                                game.phase = "setup"
                                camera_reset()
                            elif key == "lab":
                                lab = PercolationLab() if not lab.experiments else lab
                                game.phase = "lab"
                                camera_reset()
                            elif key == "brief":
                                game.phase = "brief"
                            elif key == "quit":
                                running = False

        # ─── БРИФИНГ ───
        elif game.phase == "brief":
            btns = draw_briefing(screen, anim_t)
            for ev in events:
                if ev.type == pygame.MOUSEBUTTONDOWN and ev.button == 1:
                    for key, rect in btns:
                        if rect.collidepoint(ev.pos):
                            if key == "back":
                                game.phase = "menu"
                            elif key == "go":
                                if game.day == 0 and len(game.placements) == 0:
                                    game.phase = "setup"
                                else:
                                    game.phase = "play" if game.day > 0 else "build"

        # ─── НАСТРОЙКА ИГРЫ (новая сцена!) ───
        elif game.phase == "setup":
            btns = draw_setup(screen, setup, anim_t)
            for ev in events:
                if ev.type == pygame.MOUSEBUTTONDOWN and ev.button == 1:
                    for tup in btns:
                        if len(tup) == 2:
                            key, rect = tup
                            extra = None
                        else:
                            key, rect, extra = tup
                        if rect.collidepoint(ev.pos):
                            if key == "back":
                                game.phase = "menu"
                            elif key == "go":
                                # Применяем настройки и стартуем
                                game = GameState(setup)
                                game.phase = "build"
                                camera_reset()
                            elif key == "map_type":
                                setup.map_type = extra
                            elif key == "random_seed":
                                setup.seed = random.randint(1, 100000)
                            elif key in ("density_slider", "budget_slider", "days_slider"):
                                # Стартуем drag
                                dragging_slider = (key, rect)
                                # Сразу обрабатываем клик
                                frac = max(0, min(1, (ev.pos[0] - rect.x) / rect.w))
                                if key == "density_slider":
                                    setup.density = round(0.1 + frac * 0.85, 2)
                                elif key == "budget_slider":
                                    bmin, bmax = 10_000_000, 100_000_000
                                    setup.budget = int(bmin + frac * (bmax - bmin))
                                    setup.budget = (setup.budget // 1_000_000) * 1_000_000
                                elif key == "days_slider":
                                    setup.max_days = 30 + int(frac * 90)
                            break

        # ─── ЛАБОРАТОРИЯ (новая сцена!) ───
        elif game.phase == "lab":
            btns = draw_lab(screen, lab, anim_t, hover_cell)
            for ev in events:
                if ev.type == pygame.MOUSEBUTTONDOWN and ev.button == 1:
                    for key, rect, extra in btns:
                        if rect.collidepoint(ev.pos):
                            if key == "lab_run_one":
                                lab.setup_experiment(animate=True)
                            elif key == "lab_batch_fast":
                                # 100 опытов с разной плотностью, без анимации
                                lab.start_batch(100, animate=False,
                                                density_range=(0.1, 0.95))
                                lab.setup_experiment(
                                    density=random.uniform(0.1, 0.95),
                                    animate=False,
                                    seed=random.randint(0, 1_000_000))
                            elif key == "lab_batch_animate":
                                lab.start_batch(100, animate=True,
                                                density_range=(0.1, 0.95))
                                lab.setup_experiment(
                                    density=random.uniform(0.1, 0.95),
                                    animate=True,
                                    seed=random.randint(0, 1_000_000))
                            elif key == "lab_stop":
                                lab.stop_batch()
                            elif key == "lab_clear":
                                lab.clear()
                            elif key == "lab_export":
                                if not lab.experiments:
                                    last_export_msg = "Сначала проведите эксперименты!"
                                else:
                                    path, err = export_lab(lab)
                                    if err:
                                        last_export_msg = f"Ошибка: {err}"
                                    else:
                                        last_export_msg = f"Сохранено: {os.path.basename(path)}"
                                last_export_time = anim_t
                            elif key == "lab_menu":
                                game.phase = "menu"
                            elif key == "lab_ftype":
                                lab.forest_type = extra
                            elif key in ("lab_density", "lab_wind", "lab_humid"):
                                dragging_slider = (key, rect)
                                frac = max(0, min(1, (ev.pos[0] - rect.x) / rect.w))
                                if key == "lab_density":
                                    lab.density = round(0.1 + frac * 0.85, 2)
                                elif key == "lab_wind":
                                    lab.wind_speed = round(frac * 25)
                                elif key == "lab_humid":
                                    lab.humidity = round(frac * 100)
                            break

            # Сообщение об экспорте
            if last_export_msg and (anim_t - last_export_time) < 4:
                draw_text(screen, last_export_msg, (SCREEN_W // 2, SCREEN_H - 6),
                          14, COL_ACCENT, center=True, bold=True)

        # ─── АНАЛИЗ ───
        elif game.phase == "analysis":
            btns = draw_analysis(screen, game, anim_t)
            for ev in events:
                if ev.type == pygame.MOUSEBUTTONDOWN and ev.button == 1:
                    for key, rect in btns:
                        if rect.collidepoint(ev.pos):
                            if key == "export":
                                path, err = export_game(game)
                                if err:
                                    last_export_msg = f"Ошибка: {err}"
                                else:
                                    last_export_msg = f"Сохранено: {os.path.basename(path)}"
                                last_export_time = anim_t
                            elif key == "menu":
                                game.phase = "menu"
            if last_export_msg and (anim_t - last_export_time) < 4:
                draw_text(screen, last_export_msg, (SCREEN_W // 2, SCREEN_H - 20), 14,
                          COL_ACCENT, center=True, bold=True)

        # ─── ОСНОВНАЯ ИГРА (build / play) ───
        elif game.phase in ("build", "play"):
            draw_map(screen, game, anim_t, hover_cell=hover_cell,
                     show_p=game.show_probabilities)
            draw_top_hud(screen, game)
            draw_bottom_panel(screen, game)
            panel_btns = draw_tool_panel(screen, game)

            if hover_cell:
                draw_tile_tooltip(screen, game, hover_cell[0], hover_cell[1],
                                   mouse_pos, show_p=game.show_probabilities)

            if game.phase == "build" and game.selected_tool in (TOOL_WATCHTOWER, TOOL_SENSOR):
                if hover_cell:
                    r = (8 if game.selected_tool == TOOL_WATCHTOWER else 3)
                    draw_radius(screen, hover_cell[0], hover_cell[1], r, (255, 170, 50, 120))

            if game.phase == "play" and game.plane_step == "select_target" and hover_cell:
                draw_radius(screen, hover_cell[0], hover_cell[1], 3, (100, 180, 255, 180))

            for ev in events:
                if ev.type != pygame.MOUSEBUTTONDOWN:
                    continue
                if ev.button != 1:
                    continue

                panel_hit = False
                for key, extra, rect in panel_btns:
                    if rect.collidepoint(ev.pos):
                        panel_hit = True
                        if key == "tool":
                            game.selected_tool = extra if game.selected_tool != extra else None
                            game.firebreak_anchor = None
                        elif key == "start":
                            game.start_season()
                        elif key == "nextday":
                            game.advance_day()
                        elif key == "unit":
                            p = game.placements[extra]
                            if p.kind == TOOL_BRIGADE:
                                game.selected_unit = extra if game.selected_unit != extra else None
                                game.plane_step = None
                            elif p.kind == TOOL_PLANE:
                                game.selected_unit = extra if game.selected_unit != extra else None
                                game.plane_step = "select_target" if game.selected_unit == extra else None
                        elif key == "toggle_slow":
                            game.slow_mo = not game.slow_mo
                        elif key == "toggle_prob":
                            game.show_probabilities = not game.show_probabilities
                        elif key == "menu":
                            game.phase = "menu"
                        break

                if panel_hit:
                    continue

                if hover_cell is None:
                    continue
                gx, gy = hover_cell

                if game.phase == "build" and game.selected_tool:
                    tool = game.selected_tool
                    if tool == TOOL_FIREBREAK:
                        if game.firebreak_anchor is None:
                            game.firebreak_anchor = (gx, gy)
                        else:
                            x0, y0 = game.firebreak_anchor
                            line = bresenham_line(x0, y0, gx, gy)
                            cost_each = 100_000
                            affordable = min(len(line), game.budget // cost_each)
                            for i in range(affordable):
                                x, y = line[i]
                                if game.grid.in_bounds(x, y):
                                    c = game.grid.cell(x, y)
                                    if c.state in (CELL_FOREST, CELL_EMPTY):
                                        c.state = CELL_FIREBREAK
                                        c.forest = F_NONE
                                        c.forest_original = F_NONE
                                        game.budget -= cost_each
                            game._log(f"Проложена минер. полоса ({affordable} клеток)")
                            game.firebreak_anchor = None
                    else:
                        game.place(tool, gx, gy)

                elif game.phase == "play" and game.selected_unit is not None:
                    uidx = game.selected_unit
                    p = game.placements[uidx]
                    if p.kind == TOOL_BRIGADE:
                        game.brigade_act(uidx, gx, gy)
                    elif p.kind == TOOL_PLANE and game.plane_step == "select_target":
                        game.plane_drop(uidx, gx, gy)

        pygame.display.flip()

    pygame.quit()
    sys.exit()


if __name__ == "__main__":
    main()
